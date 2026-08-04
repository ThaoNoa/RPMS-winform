# -*- coding: utf-8 -*-
"""
RPMS — Full Test Case Suite Generator (QA Lead)
Output: Docs/RPMS_TestCases.xlsx + Docs/RPMS_TestCases_Assumptions.md
Target: 400–600 detailed executable test cases
"""
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = Path(r"E:\DoAn\RPMS\Docs\RPMS_TestCases.xlsx")
ASSUMPTIONS = Path(r"E:\DoAn\RPMS\Docs\RPMS_TestCases_Assumptions.md")
OUT.parent.mkdir(parents=True, exist_ok=True)

COLS = [
    "Test Case ID", "Module", "Feature", "Requirement", "Priority",
    "Pre-condition", "Test Data", "Test Steps", "Expected Result",
    "Actual Result", "Status", "Severity", "Type", "Role", "Automation Candidate",
]

cases = []
seq = {}


def tid(module_prefix: str) -> str:
    seq[module_prefix] = seq.get(module_prefix, 0) + 1
    return f"TC-{module_prefix}-{seq[module_prefix]:03d}"


def add(
    prefix, module, feature, requirement, priority, pre, data, steps, expected,
    severity="Major", typ="Functional", role="All", auto="Yes",
):
    cases.append({
        "Test Case ID": tid(prefix),
        "Module": module,
        "Feature": feature,
        "Requirement": requirement,
        "Priority": priority,
        "Pre-condition": pre,
        "Test Data": data,
        "Test Steps": steps,
        "Expected Result": expected,
        "Actual Result": "",
        "Status": "",
        "Severity": severity,
        "Type": typ,
        "Role": role,
        "Automation Candidate": auto,
    })


# =============================================================================
# AUTH / SESSION
# =============================================================================
M = "Authentication"
add("AUTH", M, "Login", "Đăng nhập thành công với tài khoản Active", "P0",
    "DB đã seed; app mở LoginForm", "Username=namlandlord; Password=123456",
    "1. Nhập username\n2. Nhập password\n3. Bấm Đăng nhập",
    "Vào MainForm; menu theo RoleID=2 Landlord; UserSession.CurrentUser được set",
    "Critical", "Smoke", "Landlord", "Yes")
add("AUTH", M, "Login", "Đăng nhập Admin", "P0", "App mở", "admin / admin123",
    "1. Nhập admin/admin123\n2. Đăng nhập", "MainForm menu Admin (UserManagement, PostManagement…)", "Critical", "Smoke", "Admin")
add("AUTH", M, "Login", "Đăng nhập Tenant", "P0", "App mở", "tenant / 123456",
    "1. Login tenant", "Menu Tenant (Tìm phòng, HĐ, Hóa đơn…)", "Critical", "Smoke", "Tenant")
add("AUTH", M, "Login", "Đăng nhập Manager", "P0", "App mở", "manager / 123456",
    "1. Login manager", "Menu Manager (Ghi chỉ số, Quản lý sự cố)", "Critical", "Smoke", "Manager")
add("AUTH", M, "Login", "Sai mật khẩu", "P0", "User tồn tại Active", "namlandlord / wrong",
    "1. Nhập sai password\n2. Đăng nhập", "Thông báo lỗi; không vào MainForm; UserSession null", "Critical", "Negative", "Landlord")
add("AUTH", M, "Login", "Username không tồn tại", "P0", "App mở", "nouser / 123456",
    "1. Login", "Lỗi đăng nhập; không vào hệ thống", "Major", "Negative", "All")
add("AUTH", M, "Login", "Username trống", "P1", "App mở", "username rỗng",
    "1. Để trống username\n2. Đăng nhập", "Validation / lỗi; không gọi login thành công", "Major", "UI Validation", "All")
add("AUTH", M, "Login", "Password trống", "P1", "App mở", "password rỗng",
    "1. Có username, password trống\n2. Đăng nhập", "Không đăng nhập thành công", "Major", "UI Validation", "All")
add("AUTH", M, "Login", "User Inactive", "P0",
    "Admin khóa user (Status=Inactive) hoặc ToggleUserStatus", "user Inactive / đúng password",
    "1. Login bằng user Inactive", "Từ chối đăng nhập (Status != Active)", "Critical", "Negative", "All")
add("AUTH", M, "Login", "SQL Injection username", "P0", "App mở", "Username=' OR '1'='1; Password=anything",
    "1. Nhập payload SQL\n2. Đăng nhập", "Không bypass auth; login fail an toàn (EF parameterized)", "Critical", "Security", "All")
add("AUTH", M, "Login", "Special characters username", "P2", "App mở", "user!@#$%",
    "1. Login", "Fail an toàn, không crash app", "Minor", "Negative", "All")
add("AUTH", M, "Login", "Very long username/password", "P2", "App mở", "username 500 ký tự; password 500 ký tự",
    "1. Paste chuỗi dài\n2. Login", "Không crash; lỗi validation hoặc fail login", "Minor", "Boundary", "All")
add("AUTH", M, "Login", "Double click Đăng nhập", "P1", "Credentials đúng", "admin/admin123",
    "1. Double-click nhanh nút Đăng nhập", "Chỉ mở 1 MainForm; không lỗi concurrency", "Major", "Concurrency", "Admin", "No")
add("AUTH", M, "Logout", "Đăng xuất quay lại Login", "P0", "Đã login", "—",
    "1. Bấm Đăng xuất", "UserSession.Logout; hiện LoginForm; không còn truy cập menu cũ", "Critical", "Smoke", "All")
add("AUTH", M, "Session", "Không dùng session sau Logout", "P0", "Login rồi Logout", "—",
    "1. Logout\n2. Thử thao tác nếu còn form", "Không thao tác được với quyền user cũ", "Critical", "Security", "All")
add("AUTH", M, "Change Password", "Đổi mật khẩu thành công", "P1",
    "Login ProfileForm; biết mật khẩu cũ", "Old=123456; New=Abc@1234; Confirm=Abc@1234",
    "1. Hồ sơ → đổi MK\n2. Lưu\n3. Logout\n4. Login MK mới", "Login bằng MK mới thành công; MK cũ fail", "Major", "Functional", "Landlord")
add("AUTH", M, "Change Password", "Sai mật khẩu cũ", "P1", "Login Profile", "Old=sai; New=xxx; Confirm=xxx",
    "1. Đổi MK với old sai", "BadRequest; MK không đổi", "Major", "Negative", "Tenant")
add("AUTH", M, "Change Password", "Confirm không khớp", "P1", "Login Profile", "New != Confirm",
    "1. Nhập confirm khác", "Validation UI/BLL; không đổi MK", "Major", "UI Validation", "All")
add("AUTH", M, "Register", "Đăng ký tài khoản mới (nếu cho phép)", "P2",
    "Mở RegisterForm từ Login", "Username mới unique; Role Tenant; đủ field",
    "1. Điền form\n2. Đăng ký\n3. Login", "User tạo Status=Active; login được (giả định Register tạo Tenant)", "Major", "Functional", "Tenant")
add("AUTH", M, "Register", "Username trùng", "P1", "Username admin đã tồn tại", "Username=admin",
    "1. Đăng ký username trùng", "BadRequest tên đăng nhập đã tồn tại", "Major", "Negative", "All")
add("AUTH", M, "Register", "Email trùng", "P1", "Email đã dùng", "Email trùng sample",
    "1. Đăng ký email trùng", "BadRequest email đã sử dụng", "Major", "Negative", "All")

# =============================================================================
# AUTHORIZATION / MENU
# =============================================================================
M = "Authorization"
roles_menus = [
    ("Admin", "UserManagement,PostManagement,AdminReviews,ActivityLog,Backup", "LandlordHouse,TenantHome,ManagerMeter"),
    ("Landlord", "LandlordHouse,LandlordContract,LandlordAssignment", "UserManagement,ManagerMeter,TenantInvoice"),
    ("Tenant", "TenantHome,TenantContract,TenantInvoice", "LandlordHouse,UserManagement,ManagerMeter"),
    ("Manager", "ManagerMeter,ManagerMaintenance", "LandlordContract,UserManagement,TenantHome"),
]
for role, allow, deny in roles_menus:
    add("AUTHZ", M, "Menu visibility", f"Role {role} thấy đúng menu", "P0",
        f"Login {role}", f"Allowed tags: {allow}",
        "1. Login\n2. Quan sát sidebar", f"Chỉ hiện menu đúng role; có Dashboard/Notifications/Profile/Calendar",
        "Critical", "Smoke", role)
    add("AUTHZ", M, "Menu hide", f"Role {role} không thấy menu trái quyền", "P0",
        f"Login {role}", f"Denied: {deny}",
        "1. Kiểm tra sidebar không có mục trái quyền", "Không hiện menu/module của role khác",
        "Critical", "Security", role)
add("AUTHZ", M, "Reports menu", "Chỉ Admin và Landlord có Báo cáo", "P1",
    "Login lần lượt 4 role", "—",
    "1. Kiểm tra menu Báo cáo", "Role 1,2 có Reports; 3,4 không", "Major", "Functional", "All")
add("AUTHZ", M, "Chat menu", "Chỉ Landlord và Tenant có Chat", "P1",
    "Login 4 role", "—",
    "1. Kiểm tra Chat", "Role 2,3 có; 1,4 không", "Major", "Functional", "All")
add("AUTHZ", M, "Bypass UI", "Manager không tạo được HĐ qua service nếu gọi nhầm form", "P0",
    "Login Manager", "—",
    "1. Xác nhận không có menu Hợp đồng\n2. (Dev) Không resolve LandlordContractForm từ menu",
    "Không truy cập chức năng tạo HĐ từ UI Manager", "Critical", "Security", "Manager", "No")
add("AUTHZ", M, "Owner check", "Landlord A không sửa nhà Landlord B", "P0",
    "2 landlord; nhà thuộc B", "Landlord A login; HouseID của B",
    "1. A cố Update/Delete nhà của B (nếu lộ ID)", "BadRequest / không thuộc OwnerID", "Critical", "Security", "Landlord")
add("AUTHZ", M, "Assignment ownership", "Landlord chỉ gán Manager cho nhà mình", "P0",
    "Landlord A; House của B", "CreateAssignment HouseID=B",
    "1. Gán manager cho nhà người khác", "BadRequest chỉ gán nhà của mình", "Critical", "Security", "Landlord")

# =============================================================================
# ADMIN USER
# =============================================================================
M = "Admin - User Management"
add("ADM-U", M, "List users", "Hiển thị danh sách user", "P0", "Login admin", "—",
    "1. Mở Quản lý người dùng", "Grid có cột ID, Username, FullName, Role, Email, Status", "Critical", "Smoke", "Admin")
add("ADM-U", M, "Search", "Tìm theo tên/username/email", "P1", "Có nhiều user", "keyword=nam",
    "1. Nhập keyword\n2. Tìm", "Chỉ hiện user khớp", "Major", "Functional", "Admin")
add("ADM-U", M, "Search empty", "Keyword rỗng hiện all", "P2", "Đã filter", "keyword trống",
    "1. Xóa keyword\n2. Tìm", "Hiện lại toàn bộ", "Minor", "Functional", "Admin")
add("ADM-U", M, "Create user", "Tạo Manager mới", "P0", "Admin; Role Manager trong combo",
    "Username unique; Role=Manager; Password hợp lệ; Status Active",
    "1. Thêm user\n2. Chọn Role Manager\n3. Lưu", "User tạo thành công; xuất hiện trong list RoleID=4", "Critical", "Functional", "Admin")
add("ADM-U", M, "Create duplicate username", "Không trùng Username", "P0", "admin tồn tại", "Username=admin",
    "1. Tạo user username=admin", "Lỗi tên đăng nhập đã tồn tại", "Critical", "Negative", "Admin")
add("ADM-U", M, "Create duplicate email", "Không trùng Email", "P1", "Email đã dùng", "Email trùng",
    "1. Tạo với email trùng", "Lỗi email đã sử dụng", "Major", "Negative", "Admin")
add("ADM-U", M, "Update user", "Sửa FullName/Phone", "P1", "User tồn tại", "FullName mới",
    "1. Sửa\n2. Lưu", "Cập nhật DB; grid refresh", "Major", "Functional", "Admin")
add("ADM-U", M, "Toggle status", "Khóa / mở khóa user", "P0", "User Active", "—",
    "1. Bấm Khóa/Mở\n2. Login bằng user đó", "Status Inactive → login fail; Active lại → login OK", "Critical", "Functional", "Admin")
add("ADM-U", M, "Delete user", "Xóa mềm / Inactive", "P1", "User không phải đang login", "—",
    "1. Xóa user", "Status Inactive (soft delete theo code)", "Major", "Functional", "Admin")
add("ADM-U", M, "Required fields", "Thiếu FullName/Username/Password", "P1", "Modal thêm", "Field bắt buộc trống",
    "1. Lưu form thiếu field", "Validation; không lưu", "Major", "UI Validation", "Admin")
add("ADM-U", M, "XSS FullName", "FullName chứa script", "P1", "Modal", "FullName=<script>alert(1)</script>",
    "1. Lưu\n2. Xem lại list", "Lưu plain text; WinForms không execute XSS web (ghi nhận)", "Minor", "Security", "Admin", "No")

# =============================================================================
# ADMIN POST
# =============================================================================
M = "Admin - Post Management"
add("ADM-P", M, "List pending", "Xem tin chờ duyệt", "P0", "Có Post Status=Pending", "—",
    "1. Mở Quản lý tin đăng", "Thấy tin Pending", "Critical", "Smoke", "Admin")
add("ADM-P", M, "Approve", "Duyệt tin", "P0", "Post Pending", "PostID sample",
    "1. Duyệt tin", "Status=Approved; ApprovedBy/Date set; Tenant search thấy tin", "Critical", "Functional", "Admin")
add("ADM-P", M, "Reject", "Từ chối tin", "P0", "Post Pending", "—",
    "1. Từ chối", "Status=Rejected; không hiện ở tìm phòng Active", "Critical", "Functional", "Admin")
add("ADM-P", M, "Filter status", "Lọc theo trạng thái", "P1", "Nhiều status", "Pending/Approved",
    "1. Đổi filter\n2. Refresh", "Grid đúng status", "Major", "Functional", "Admin")
add("ADM-P", M, "Approve twice", "Duyệt lại tin đã Approved", "P2", "Post Approved", "—",
    "1. Thử duyệt lại", "Không lỗi nghiêm trọng / hoặc no-op / thông báo", "Minor", "Negative", "Admin")

# =============================================================================
# ADMIN OTHER
# =============================================================================
M = "Admin - Reviews / ActivityLog / Backup"
add("ADM-O", M, "Reviews list", "Admin xem mọi đánh giá", "P1", "Có Review", "—",
    "1. Mở Đánh giá", "Grid reviews", "Major", "Functional", "Admin")
add("ADM-O", M, "Activity log", "Xem nhật ký", "P1", "Đã có login log", "—",
    "1. Mở Nhật ký\n2. Làm mới", "Thấy Action đăng nhập gần đây", "Major", "Functional", "Admin")
add("ADM-O", M, "Backup", "Backup DB (nếu form có)", "P2", "SQL quyền backup; form Backup", "Đường dẫn hợp lệ",
    "1. Chạy backup", "File backup tạo được hoặc thông báo lỗi rõ nếu thiếu implement", "Major", "Functional", "Admin", "No")

# =============================================================================
# LANDLORD HOUSE
# =============================================================================
M = "Landlord - House"
add("LL-H", M, "List houses", "Danh sách nhà của owner", "P0", "Login namlandlord", "OwnerID=2",
    "1. Mở Nhà của tôi", "Chỉ nhà OwnerID=current user", "Critical", "Smoke", "Landlord")
add("LL-H", M, "Create house", "Tạo nhà mới", "P0", "Login landlord", "HouseName, Address bắt buộc",
    "1. Thêm nhà\n2. Lưu", "House Status=Active; OwnerID=landlord; hiện list", "Critical", "Functional", "Landlord")
add("LL-H", M, "Create missing address", "Thiếu Address", "P1", "Modal", "Address trống",
    "1. Lưu", "Validation fail", "Major", "UI Validation", "Landlord")
add("LL-H", M, "Update house", "Sửa tên/địa chỉ", "P1", "Có nhà", "HouseName mới",
    "1. Sửa\n2. Lưu", "UpdatedDate đổi; data mới", "Major", "Functional", "Landlord")
add("LL-H", M, "Delete/Inactive house", "Xóa hoặc ngưng nhà", "P1", "Có nhà", "—",
    "1. Xóa", "Theo implement: xóa hoặc Inactive; không còn thao tác bình thường", "Major", "Functional", "Landlord")
add("LL-H", M, "Long address", "Address rất dài", "P2", "Modal", "Address > 255 ký tự",
    "1. Lưu", "Lỗi độ dài hoặc truncate an toàn; không crash", "Minor", "Boundary", "Landlord")

# =============================================================================
# LANDLORD ROOM
# =============================================================================
M = "Landlord - Room"
add("LL-R", M, "List rooms", "Phòng theo nhà", "P0", "Có nhà + phòng", "HouseID",
    "1. Chọn nhà\n2. Xem grid", "Rooms đúng HouseID", "Critical", "Smoke", "Landlord")
add("LL-R", M, "Create room", "Tạo phòng Available", "P0", "Có nhà", "RoomNumber unique trong nhà; Price>0; Area>0",
    "1. Thêm phòng\n2. Lưu", "Status=Available; UQ(HouseID,RoomNumber) OK", "Critical", "Functional", "Landlord")
add("LL-R", M, "Duplicate room number", "Trùng số phòng cùng nhà", "P0", "Đã có phòng 101", "RoomNumber=101 cùng House",
    "1. Tạo trùng", "Lỗi Unique / BadRequest", "Critical", "Negative", "Landlord")
add("LL-R", M, "Price zero", "Price <= 0", "P0", "Modal", "Price=0",
    "1. Lưu", "CHECK Price>0 hoặc validation UI", "Critical", "Boundary", "Landlord")
add("LL-R", M, "Area zero", "Area <= 0", "P1", "Modal", "Area=0",
    "1. Lưu", "Validation fail", "Major", "Boundary", "Landlord")
add("LL-R", M, "Capacity boundary", "Capacity=0", "P1", "Modal", "Capacity=0",
    "1. Lưu", "CHECK Capacity>=1 fail", "Major", "Boundary", "Landlord")
add("LL-R", M, "Upload images", "Upload nhiều ảnh", "P1", "Phòng tồn tại; file jpg/png", "2–5 ảnh",
    "1. Upload\n2. Lưu\n3. Xem chi tiết/Tenant", "Ảnh lưu path; gallery hiển thị", "Major", "Functional", "Landlord", "No")
add("LL-R", M, "Upload invalid file", "File không phải ảnh/video", "P2", "Modal", "file .exe",
    "1. Chọn file lạ", "Từ chối hoặc không preview; không crash", "Minor", "Negative", "Landlord", "No")
add("LL-R", M, "Assign amenities", "Gán tiện nghi", "P1", "Có Amenities seed", "Wifi, AC…",
    "1. Chọn amenities\n2. Lưu", "RoomAmenities ghi đúng", "Major", "Functional", "Landlord")
add("LL-R", M, "Cannot edit occupied blindly", "Đổi status Occupied thủ công khi có HĐ", "P1",
    "Phòng Occupied có HĐ Active", "—",
    "1. Quan sát/ cập nhật status", "Status đồng bộ với HĐ; không để Available nếu còn Active", "Major", "Business Rule", "Landlord")

# =============================================================================
# LANDLORD ASSIGNMENT
# =============================================================================
M = "Landlord - Assignment"
add("LL-A", M, "Find manager by ID", "Tìm Manager ID=4", "P0", "Login landlord; có nhà", "Manager ID=4",
    "1. Nhập 4\n2. Tìm", "Preview ✓ tên Manager Active", "Critical", "Smoke", "Landlord")
add("LL-A", M, "Find by username", "Tìm username=manager", "P0", "Có manager", "manager",
    "1. Nhập manager\n2. Tìm", "Tìm thấy Role Manager", "Critical", "Functional", "Landlord")
add("LL-A", M, "Find non-manager", "ID Tenant không phải Manager", "P0", "Tenant ID=3", "3",
    "1. Tìm ID=3", "Báo không phải Manager", "Critical", "Negative", "Landlord")
add("LL-A", M, "Assign success", "Gán Manager cho nhà", "P0", "Manager mới hoặc chưa Active pair",
    "House của landlord; Manager Active Role=4",
    "1. Tìm manager\n2. Chọn nhà\n3. Gán", "Assignment Active; notify Manager; hiện grid", "Critical", "Functional", "Landlord")
add("LL-A", M, "Assign duplicate active", "Gán trùng Active cùng House+Manager", "P0",
    "Đã có Assignment Active (1,4)", "House=1 Manager=4",
    "1. Gán lại", "BadRequest đã gán Active", "Critical", "Negative", "Landlord")
add("LL-A", M, "Reactivate inactive", "Gán lại sau khi Ngưng", "P1",
    "Deactivate assignment trước", "Cùng House+Manager",
    "1. Ngưng\n2. Gán lại", "Reactivate Status=Active; AssignedDate mới; notify", "Major", "Functional", "Landlord")
add("LL-A", M, "Deactivate", "Ngưng phân công", "P0", "Có Active assignment", "—",
    "1. Bấm Ngưng\n2. Confirm", "Status=Inactive; Manager không còn thấy HĐ nhà đó", "Critical", "Functional", "Landlord")
add("LL-A", M, "Inactive manager", "Manager Status Inactive", "P1", "Admin khóa manager", "Manager Inactive",
    "1. Tìm/Gán", "Không cho gán / báo không Active", "Major", "Negative", "Landlord")
add("LL-A", M, "No house selected", "Chưa chọn nhà", "P1", "Form mở", "Không chọn nhà",
    "1. Gán", "Warning chọn nhà", "Major", "UI Validation", "Landlord")
add("LL-A", M, "No find before assign", "Chưa Tìm manager", "P1", "Form", "_foundManager null",
    "1. Bấm Gán ngay", "Warning hãy tìm Manager trước", "Major", "UI Validation", "Landlord")
add("LL-A", M, "UI usable resize", "Nút Gán không bị cắt khi thu nhỏ", "P1", "Form Assignment", "Resize hẹp",
    "1. Thu nhỏ cửa sổ", "Nút Tìm/Gán vẫn bấm được (FlowLayout)", "Major", "Usability", "Landlord", "No")

# =============================================================================
# LANDLORD CONTRACT
# =============================================================================
M = "Landlord - Contract"
add("LL-C", M, "List contracts", "Danh sách HĐ của landlord", "P0", "Login landlord; có HĐ", "—",
    "1. Mở Hợp đồng", "Grid HĐ thuộc nhà owner", "Critical", "Smoke", "Landlord")
add("LL-C", M, "Create with tenant", "Tạo HĐ Active có khách", "P0",
    "Phòng Available; tenant đã đặt lịch", "Room Available; Tenant từ combo; dates End>Start; Rent>0",
    "1. Chọn nhà/phòng/khách\n2. Nhập tiền\n3. Lưu",
    "Status=Active; Room=Occupied; Tenant nhận Notification; ContractCode HD…", "Critical", "Functional", "Landlord")
add("LL-C", M, "Create draft no tenant", "Lưu nháp chưa khách", "P0", "Phòng Available", "Tenant=(Chưa có khách)",
    "1. Để trống khách\n2. Lưu", "Status=Draft; Room vẫn Available; chưa notify tenant", "Critical", "Functional", "Landlord")
add("LL-C", M, "Room occupied", "Không tạo HĐ phòng Occupied", "P0", "Phòng Occupied", "Room Occupied",
    "1. Cố tạo (nếu lộ RoomID)", "BadRequest đã có người thuê", "Critical", "Business Rule", "Landlord")
add("LL-C", M, "Existing Active/Draft", "Không 2 HĐ mở cùng phòng", "P0", "Phòng đã Draft hoặc Active", "Cùng RoomID",
    "1. Tạo thêm", "BadRequest đã có HĐ nháp hoặc hiệu lực", "Critical", "Business Rule", "Landlord")
add("LL-C", M, "EndDate <= StartDate", "Ngày kết thúc không hợp lệ", "P0", "Form tạo", "End <= Start",
    "1. Lưu", "BadRequest ngày kết thúc phải lớn hơn bắt đầu", "Critical", "Validation", "Landlord")
add("LL-C", M, "Rent invalid", "Tiền thuê <=0", "P1", "Form", "Rent=0",
    "1. Lưu", "Warning số tiền hợp lệ", "Major", "UI Validation", "Landlord")
add("LL-C", M, "Bulk draft", "Tạo nháp tất cả phòng trống", "P0",
    "Nhà có ≥2 phòng Available chưa HĐ", "Deposit/Electric/Water hợp lệ; Rent trống → lấy giá phòng",
    "1. Chọn nhà\n2. Tạo nháp tất cả\n3. Confirm",
    "CreatedCount>0; mỗi phòng 1 Draft; phòng đã có HĐ bị skip", "Critical", "Functional", "Landlord")
add("LL-C", M, "Bulk none eligible", "Không còn phòng để bulk", "P1", "Mọi phòng đã có HĐ hoặc Occupied", "—",
    "1. Bulk", "Message không còn phòng; CreatedCount=0", "Major", "Negative", "Landlord")
add("LL-C", M, "Assign tenant to draft", "Gán khách lên HĐ nháp", "P0", "HĐ Draft; tenant appointment", "—",
    "1. Gán khách trên grid", "Active; Occupied; notify tenant", "Critical", "Functional", "Landlord")
add("LL-C", M, "Assign when already has tenant", "Gán lại khi đã có khách", "P1", "HĐ Active có TenantID", "—",
    "1. Gán khách", "BadRequest đã có khách thuê", "Major", "Negative", "Landlord")
add("LL-C", M, "Edit draft immediate", "Sửa HĐ Draft áp dụng ngay", "P0", "HĐ Draft", "MonthlyRent mới",
    "1. Sửa\n2. Lưu", "Giá mới áp dụng ngay; không Pending", "Critical", "Functional", "Landlord")
add("LL-C", M, "Edit active pending", "Sửa HĐ Active → Pending tenant confirm", "P0", "HĐ Active có tenant", "Rent mới",
    "1. Sửa\n2. Lưu", "PendingEditStatus=Pending; notify tenant; giá cũ còn hiệu lực đến khi confirm", "Critical", "Workflow", "Landlord")
add("LL-C", M, "Cancel pending edit", "Chủ hủy đề xuất", "P1", "Đang Pending", "—",
    "1. Hủy đề xuất", "Pending cleared; giá không đổi", "Major", "Functional", "Landlord")
add("LL-C", M, "Extend contract", "Gia hạn EndDate", "P1", "HĐ Active", "NewEnd > End cũ",
    "1. Gia hạn", "EndDate mới; notify", "Major", "Functional", "Landlord")
add("LL-C", M, "Terminate", "Hủy/chấm dứt HĐ", "P1", "HĐ Active", "—",
    "1. Hủy HĐ", "Status Terminated/Expired theo code; Room Available", "Major", "Functional", "Landlord")
add("LL-C", M, "Print PDF", "In/PDF hợp đồng", "P2", "Có HĐ", "—",
    "1. In/PDF", "Mở preview/HTML print được", "Minor", "Functional", "Landlord", "No")
add("LL-C", M, "Tenant combo only bookers", "Combo khách chỉ người đặt lịch", "P1",
    "Tenant chưa đặt lịch vs đã đặt", "—",
    "1. Mở combo khách theo phòng", "Chỉ GetAppointmentTenantsAsync", "Major", "Business Rule", "Landlord")

# =============================================================================
# LANDLORD APPOINTMENT / POST / REVIEW
# =============================================================================
M = "Landlord - Appointment"
add("LL-AP", M, "List appointments", "Lọc lịch hẹn", "P0", "Có appointment Pending", "House/Status/Date",
    "1. Mở Lịch hẹn\n2. Lọc", "Grid có RoomNumber, TenantName, Status", "Critical", "Smoke", "Landlord")
add("LL-AP", M, "Accept", "Xác nhận lịch → notify tenant", "P0", "Appointment Pending", "—",
    "1. Bấm Nhận\n2. Login tenant xem Thông báo",
    "Status=Accepted; Notification tenant tiêu đề xác nhận; nội dung tiếng Việt có phòng/giờ", "Critical", "Functional", "Landlord")
add("LL-AP", M, "Reject", "Từ chối + notify", "P0", "Pending", "—",
    "1. Từ chối\n2. Tenant xem notify", "Status=Rejected; notify từ chối", "Critical", "Functional", "Landlord")
add("LL-AP", M, "Complete", "Hoàn thành lịch", "P1", "Accepted", "—",
    "1. Xong", "Status=Completed (DB cho phép); notify", "Major", "Functional", "Landlord")
add("LL-AP", M, "Filter date range", "Lọc khoảng ngày", "P1", "Nhiều lịch", "From/To",
    "1. Đổi ngày\n2. Lọc", "Chỉ lịch trong khoảng", "Major", "Functional", "Landlord")
add("LL-AP", M, "Invalid status", "Set status không hợp lệ", "P2", "Dev/API", "Status=Foo",
    "1. Gọi Update với status lạ", "BadRequest trạng thái không hợp lệ", "Minor", "Negative", "Landlord", "Yes")

M = "Landlord - Post"
add("LL-PO", M, "Create post", "Đăng tin Pending", "P0", "Có phòng", "Title, PriceSnapshot, Room",
    "1. Đăng tin\n2. Lưu", "Status=Pending; chờ Admin duyệt", "Critical", "Functional", "Landlord")
add("LL-PO", M, "Missing title", "Thiếu tiêu đề", "P1", "Form", "Title trống",
    "1. Lưu", "Validation", "Major", "UI Validation", "Landlord")
add("LL-PO", M, "Featured flag", "Đánh dấu nổi bật", "P2", "Form", "IsFeatured=true",
    "1. Lưu\n2. Admin duyệt\n3. Tenant filter nổi bật", "Tin xuất hiện khi FeaturedOnly", "Minor", "Functional", "Landlord")

M = "Landlord - Review"
add("LL-RV", M, "Reply review", "Chủ trả lời đánh giá", "P1", "Có Review của tenant", "Reply text",
    "1. Reply\n2. Lưu", "LandlordReply + Date; tenant thấy", "Major", "Functional", "Landlord")
add("LL-RV", M, "Reply empty", "Reply trống", "P2", "Form", "Reply rỗng",
    "1. Lưu", "Validation hoặc cho phép rỗng theo rule", "Minor", "UI Validation", "Landlord")

# =============================================================================
# TENANT SEARCH / FAVORITE / APPOINTMENT
# =============================================================================
M = "Tenant - Search & Room"
add("TN-S", M, "Search default", "Tìm phòng còn trống", "P0", "Login tenant; có Post Approved", "Status=Còn trống",
    "1. Mở Tìm phòng\n2. Tìm kiếm", "Hiện card phòng Available/Approved", "Critical", "Smoke", "Tenant")
add("TN-S", M, "Filter price range", "Lọc giá từ-đến", "P1", "Nhiều giá", "Min=2000000 Max=5000000",
    "1. Nhập giá\n2. Tìm", "Chỉ phòng trong khoảng", "Major", "Functional", "Tenant")
add("TN-S", M, "Filter price min>max", "Giá từ > đến", "P1", "Form", "Min>Max",
    "1. Tìm", "Không kết quả hoặc validation", "Major", "Negative", "Tenant")
add("TN-S", M, "Filter area", "Lọc diện tích", "P1", "Combo diện tích", "25–50m²",
    "1. Chọn\n2. Tìm", "Đúng bucket diện tích", "Major", "Functional", "Tenant")
add("TN-S", M, "Filter amenities", "Lọc tiện nghi", "P1", "Có phòng wifi", "chkWifi",
    "1. Tick Wifi\n2. Tìm", "Chỉ phòng có amenity", "Major", "Functional", "Tenant")
add("TN-S", M, "Filter featured", "Chỉ tin nổi bật", "P1", "Có featured", "chkFeatured",
    "1. Tick\n2. Tìm", "Chỉ IsFeatured", "Major", "Functional", "Tenant")
add("TN-S", M, "Sort price", "Sắp xếp giá tăng/giảm", "P1", "Nhiều kết quả", "Giá tăng",
    "1. Chọn sort\n2. Tìm", "Thứ tự đúng", "Major", "Functional", "Tenant")
add("TN-S", M, "Clear filters", "Xóa lọc", "P2", "Đã lọc", "—",
    "1. Xóa lọc", "Reset control; tìm lại mặc định", "Minor", "Functional", "Tenant")
add("TN-S", M, "Empty result", "Không có kết quả", "P1", "Filter quá hẹp", "Giá cực lớn",
    "1. Tìm", "Empty state thân thiện", "Major", "Usability", "Tenant")
add("TN-S", M, "Room detail", "Xem chi tiết + gallery", "P0", "Có ảnh", "—",
    "1. Click card/chi tiết", "RoomDetailForm; ảnh/video", "Critical", "Functional", "Tenant", "No")
add("TN-S", M, "Search special chars", "Keyword ký tự đặc biệt", "P2", "Form", "%%%___'",
    "1. Tìm", "Không SQL error; kết quả rỗng hoặc escape", "Minor", "Security", "Tenant")
add("TN-S", M, "UI wrap filters", "Filter không đè khi resize", "P1", "TenantHome", "Thu nhỏ cửa sổ",
    "1. Resize", "FlowLayout wrap; vẫn dùng được", "Major", "Usability", "Tenant", "No")

M = "Tenant - Favorite"
add("TN-F", M, "Toggle favorite", "Thêm yêu thích", "P0", "Login tenant; có phòng", "RoomID",
    "1. Yêu thích\n2. Mở danh sách yêu thích", "Có trong Favorites; Unique User+Room", "Critical", "Functional", "Tenant")
add("TN-F", M, "Remove favorite", "Bỏ yêu thích", "P1", "Đã favorite", "—",
    "1. Bỏ/Xóa", "Không còn trong list", "Major", "Functional", "Tenant")
add("TN-F", M, "Duplicate favorite", "Favorite 2 lần", "P1", "Đã favorite", "Cùng Room",
    "1. Toggle lại", "Idempotent: bỏ hoặc báo đã có (UQ)", "Major", "Negative", "Tenant")

M = "Tenant - Appointment"
add("TN-AP", M, "Book appointment", "Đặt lịch xem phòng", "P0", "Post Approved; phòng Available",
    "AppointmentDate tương lai; Note",
    "1. Đặt lịch\n2. Landlord xem", "Status=Pending; notify Landlord", "Critical", "Functional", "Tenant")
add("TN-AP", M, "Past date", "Ngày hẹn quá khứ", "P1", "Modal", "Date < Now",
    "1. Đặt", "Validation không cho / hoặc chấp nhận theo rule (ghi giả định)", "Major", "Boundary", "Tenant")
add("TN-AP", M, "Note long", "Note rất dài", "P2", "Modal", "Note 5000 ký tự",
    "1. Đặt", "Lỗi max length hoặc cắt; không crash", "Minor", "Boundary", "Tenant")

# =============================================================================
# TENANT CONTRACT / INVOICE / MAINTENANCE / REVIEW
# =============================================================================
M = "Tenant - Contract"
add("TN-C", M, "View contracts", "Xem HĐ của tôi", "P0", "Tenant có HĐ", "TenantID=3",
    "1. Mở Hợp đồng của tôi", "Chỉ HĐ TenantID=current", "Critical", "Smoke", "Tenant")
add("TN-C", M, "Confirm pending edit", "Xác nhận đề xuất sửa giá", "P0",
    "Landlord đã Pending edit", "—",
    "1. Xem đề xuất\n2. Xác nhận", "Giá mới áp dụng; Previous* lưu; PriceEffectiveDate; Pending clear", "Critical", "Workflow", "Tenant")
add("TN-C", M, "Reject pending edit", "Từ chối đề xuất", "P0", "Pending", "—",
    "1. Từ chối", "Giá cũ giữ; Pending clear; landlord biết (notify nếu có)", "Critical", "Workflow", "Tenant")
add("TN-C", M, "Confirm without pending", "Xác nhận khi không Pending", "P1", "HĐ không Pending", "—",
    "1. Thử confirm", "Lỗi / nút disable", "Major", "Negative", "Tenant")
add("TN-C", M, "Other tenant contract", "Không xem HĐ người khác", "P0", "HĐ của tenant khác", "—",
    "1. Chỉ thấy HĐ mình trên UI", "Không lộ HĐ khác", "Critical", "Security", "Tenant")

M = "Tenant - Invoice"
add("TN-I", M, "List invoices", "Xem hóa đơn", "P0", "Có Invoice", "—",
    "1. Mở Hóa đơn", "List Unpaid/Paid", "Critical", "Smoke", "Tenant")
add("TN-I", M, "Detail", "Chi tiết hóa đơn đủ block", "P0", "Có invoice", "—",
    "1. Xem chi tiết", "Thông tin HĐ/phòng/khách; điện nước; tổng", "Critical", "Functional", "Tenant")
add("TN-I", M, "Pay", "Thanh toán Unpaid", "P0", "Invoice Unpaid", "Method Banking",
    "1. Thanh toán", "Status=Paid; Payment record; PaidDate", "Critical", "Functional", "Tenant")
add("TN-I", M, "Pay twice", "Thanh toán lại hóa đơn Paid", "P0", "Invoice Paid", "—",
    "1. Thanh toán lại", "Không cho / BadRequest", "Critical", "Negative", "Tenant")
add("TN-I", M, "Export excel", "Xuất CSV/Excel list", "P2", "Có data", "—",
    "1. Xuất Excel", "File CSV tạo được", "Minor", "Functional", "Tenant", "No")
add("TN-I", M, "Print PDF", "In PDF chi tiết", "P2", "Detail form", "—",
    "1. In/PDF", "Preview OK", "Minor", "Functional", "Tenant", "No")

M = "Tenant - Maintenance"
add("TN-M", M, "Create request", "Báo sự cố", "P0", "Có HĐ Active", "Title, Description; ảnh optional",
    "1. Gửi yêu cầu", "Status=Pending; Manager/Landlord thấy theo nhà", "Critical", "Functional", "Tenant")
add("TN-M", M, "Missing title", "Thiếu tiêu đề", "P1", "Form", "Title trống",
    "1. Gửi", "Validation", "Major", "UI Validation", "Tenant")
add("TN-M", M, "No active contract", "Không có HĐ Active", "P1", "Tenant không HĐ", "—",
    "1. Mở form / gửi", "Không chọn được contract hoặc lỗi", "Major", "Negative", "Tenant")
add("TN-M", M, "Upload image", "Đính kèm ảnh sự cố", "P1", "File ảnh", "jpg",
    "1. Upload\n2. Gửi\n3. Manager xem detail", "Ảnh hiển thị", "Major", "Functional", "Tenant", "No")

M = "Tenant - Review"
add("TN-RV", M, "Create review", "Đánh giá sau thuê", "P0", "Có HĐ; chưa review", "Rating 1–5; Comment",
    "1. Chọn HĐ\n2. Gửi đánh giá", "Review tạo; UQ theo Contract; landlord thấy", "Critical", "Functional", "Tenant")
add("TN-RV", M, "Duplicate review", "Review 2 lần cùng HĐ", "P0", "Đã review", "Cùng ContractID",
    "1. Gửi lại", "Lỗi unique / không cho", "Critical", "Negative", "Tenant")
add("TN-RV", M, "Rating boundary", "Rating=0 hoặc 6", "P1", "Form", "Rating ngoài 1–5",
    "1. Gửi", "NumericUpDown chặn hoặc CHECK fail", "Major", "Boundary", "Tenant")
add("TN-RV", M, "Rating=1 and 5", "Biên hợp lệ", "P1", "Form", "1 và 5",
    "1. Gửi từng giá trị", "Cả hai thành công", "Major", "Boundary", "Tenant")

# =============================================================================
# MANAGER METER / MAINTENANCE
# =============================================================================
M = "Manager - Meter & Invoice"
add("MG-M", M, "List active contracts", "Thấy HĐ Active có khách nhà được gán", "P0",
    "Assignment Active House1; HĐ Active", "manager login",
    "1. Mở Ghi chỉ số", "Grid có HD Active+Tenant; Draft không hiện", "Critical", "Smoke", "Manager")
add("MG-M", M, "No assignment", "Chưa được gán nhà", "P0", "Manager mới không Assignment", "—",
    "1. Mở form", "Empty state chưa phân công", "Critical", "Functional", "Manager")
add("MG-M", M, "After assign", "Gán xong thấy HĐ", "P0", "Landlord vừa gán nhà có HĐ", "—",
    "1. Login manager\n2. Làm mới", "Thấy HĐ nhà mới gán", "Critical", "Regression", "Manager")
add("MG-M", M, "Generate invoice prev month", "Tạo HĐ tháng trước", "P0",
    "Chưa có reading tháng T-1; HĐ Active có ngày ở", "NewElectric>=Old; NewWater>=Old; OtherFee>=0",
    "1. Chọn HĐ\n2. Nhập chỉ số\n3. Tạo hóa đơn",
    "MeterReading + Invoice Unpaid; notify tenant; BillingMonth=tháng trước", "Critical", "Functional", "Manager")
add("MG-M", M, "Duplicate month", "Tạo trùng tháng", "P0", "Đã có reading tháng T-1", "Cùng tháng",
    "1. Tạo lại", "BadRequest đã có chỉ số/hóa đơn tháng …", "Critical", "Business Rule", "Manager")
add("MG-M", M, "Current month blocked", "Không tạo tháng hiện tại", "P0", "Hôm nay trong tháng", "ReadingMonth=tháng này",
    "1. (UI luôn tháng trước) nếu force", "BadRequest chỉ tháng đã kết thúc", "Critical", "Business Rule", "Manager")
add("MG-M", M, "New < Old electric", "Chỉ số điện giảm", "P0", "Có prev", "NewElectric < Old",
    "1. Nhập\n2. Tạo", "Warning UI và/hoặc BadRequest", "Critical", "Validation", "Manager")
add("MG-M", M, "New < Old water", "Chỉ số nước giảm", "P0", "Có prev", "NewWater < Old",
    "1. Tạo", "Lỗi validation", "Critical", "Validation", "Manager")
add("MG-M", M, "First reading", "Lần đầu Old=0", "P1", "HĐ mới chưa reading", "New>0",
    "1. Tạo", "Old=0; hóa đơn OK nếu có ngày ở", "Major", "Functional", "Manager")
add("MG-M", M, "Prorate mid-month move-in", "Tiền nhà prorate", "P0",
    "MoveIn ngày 15; tạo hóa đơn tháng đó (nếu còn trong rule tháng trước)", "—",
    "1. Tạo hóa đơn tháng có move-in giữa tháng", "Rent < full month; OccupiedDays đúng", "Critical", "Business Rule", "Manager")
add("MG-M", M, "Price change mid-month", "Đổi giá sau tenant confirm", "P1",
    "Pending confirmed với PriceEffectiveDate giữa tháng", "—",
    "1. Tạo hóa đơn tháng đó", "WeightedUnitCost / CalculateRent dùng Previous+New", "Critical", "Business Rule", "Manager")
add("MG-M", M, "Non-numeric input", "Nhập chữ vào chỉ số", "P1", "Form", "abc",
    "1. Tạo", "Warning nhập số hợp lệ", "Major", "UI Validation", "Manager")
add("MG-M", M, "Double submit", "Double click tạo HĐ", "P1", "Chưa có invoice tháng", "—",
    "1. Double-click nhanh", "Chỉ 1 invoice; lần 2 báo đã có", "Major", "Concurrency", "Manager", "No")
add("MG-M", M, "Draft not listed", "HĐ Draft không ghi được", "P1", "Chỉ Draft trên nhà", "—",
    "1. Xem grid", "Không có Draft; empty hint Active+khách", "Major", "Business Rule", "Manager")

M = "Manager - Maintenance"
add("MG-MT", M, "List requests", "Xem sự cố nhà được gán", "P0", "Assignment + request", "—",
    "1. Mở Quản lý sự cố", "Chỉ request thuộc house được gán", "Critical", "Smoke", "Manager")
add("MG-MT", M, "Accept processing", "Tiếp nhận → Processing", "P0", "Request Pending", "—",
    "1. Xác nhận & Hẹn", "Status=Processing; notify tenant", "Critical", "Workflow", "Manager")
add("MG-MT", M, "Complete", "Hoàn thành", "P0", "Processing", "—",
    "1. Xong", "Status=Completed; CompletedDate; notify", "Critical", "Workflow", "Manager")
add("MG-MT", M, "Detail timeline", "Xem timeline trạng thái", "P1", "Detail form", "—",
    "1. Xem chi tiết", "StatusTimeline hiển thị các mốc", "Major", "Usability", "Manager", "No")
add("MG-MT", M, "Print", "In/PDF phiếu", "P2", "Detail", "—",
    "1. In", "Preview OK", "Minor", "Functional", "Manager", "No")
add("MG-MT", M, "Other house request", "Không thấy sự cố nhà không gán", "P0",
    "Request nhà khác", "—",
    "1. So sánh list", "Không xuất hiện", "Critical", "Security", "Manager")

# =============================================================================
# NOTIFICATION / PROFILE / CHAT / CALENDAR / REPORT / DASHBOARD
# =============================================================================
M = "Shared - Notification"
add("SH-N", M, "List notifications", "Trung tâm thông báo", "P0", "Có notify", "Login user nhận",
    "1. Mở Thông báo\n2. Làm mới", "Thấy Title/Content/CreatedDate; unread count", "Critical", "Smoke", "All")
add("SH-N", M, "Mark read", "Đánh dấu đã đọc", "P1", "Unread", "—",
    "1. Đánh dấu đọc", "IsRead=true; count giảm", "Major", "Functional", "All")
add("SH-N", M, "Mark all", "Đọc tất cả", "P1", "Nhiều unread", "—",
    "1. Đọc tất cả", "Tất cả IsRead", "Major", "Functional", "All")
add("SH-N", M, "Delete", "Xóa thông báo", "P2", "Có item", "—",
    "1. Xóa confirm", "Mất khỏi list", "Minor", "Functional", "All")
add("SH-N", M, "Filter unread", "Lọc chưa đọc", "P1", "Mix read/unread", "Filter Chưa đọc",
    "1. Lọc", "Chỉ IsRead=false", "Major", "Functional", "All")
add("SH-N", M, "Search keyword", "Tìm nội dung", "P2", "Có keyword", "từ khóa trong Title",
    "1. Tìm", "Khớp contains", "Minor", "Functional", "All")
add("SH-N", M, "Isolation", "Không thấy notify user khác", "P0", "2 user", "—",
    "1. So sánh UserID", "Chỉ notify của CurrentUser", "Critical", "Security", "All")
add("SH-N", M, "Refresh activated", "Activated reload", "P1", "Form đã mở; có notify mới", "—",
    "1. Chuyển form khác rồi vào lại", "List cập nhật", "Major", "Functional", "All", "No")

M = "Shared - Profile"
add("SH-P", M, "View profile", "Xem hồ sơ", "P0", "Login", "—",
    "1. Mở Hồ sơ", "FullName Email Phone Address đúng", "Critical", "Smoke", "All")
add("SH-P", M, "Update profile", "Cập nhật thông tin", "P1", "Profile", "Phone mới",
    "1. Sửa\n2. Lưu", "DB cập nhật", "Major", "Functional", "All")
add("SH-P", M, "Activity logs on profile", "Xem log cá nhân", "P2", "Có activity", "—",
    "1. Xem grid log", "Log của user", "Minor", "Functional", "All")
add("SH-P", M, "Splitter resize", "Form Profile không crash khi mở", "P0", "Login", "—",
    "1. Mở Profile", "Không lỗi SplitterDistance; UI dùng được", "Critical", "Regression", "All")

M = "Shared - Chat"
add("SH-C", M, "Open chat", "Landlord-Tenant chat", "P0", "Login Landlord hoặc Tenant", "—",
    "1. Mở Chat\n2. Chọn/tạo conversation", "Gửi nhận tin được", "Critical", "Functional", "Landlord")
add("SH-C", M, "Send message", "Gửi text", "P0", "Có conversation", "Content ngắn",
    "1. Gửi", "Message lưu; đối phương thấy; LastMessageAt", "Critical", "Functional", "Tenant")
add("SH-C", M, "Empty message", "Gửi rỗng", "P1", "Chat", "Content trống",
    "1. Gửi", "Không gửi / validation", "Major", "UI Validation", "Landlord")
add("SH-C", M, "Mark read", "Đánh dấu đã đọc", "P1", "Có unread", "—",
    "1. Mở conversation", "Unread giảm", "Major", "Functional", "Tenant")
add("SH-C", M, "Admin no chat menu", "Admin không chat", "P1", "Login admin", "—",
    "1. Kiểm tra menu", "Không có Chat", "Major", "Authorization", "Admin")

M = "Shared - Calendar / Report / Dashboard"
add("SH-D", M, "Dashboard open", "Dashboard không crash", "P0", "Login bất kỳ role", "—",
    "1. Mở Dashboard (mặc định)", "Load stats; không Sequence contains no elements", "Critical", "Smoke", "All")
add("SH-D", M, "Dashboard landlord cards", "Thẻ thống kê landlord", "P1", "Login landlord", "—",
    "1. Xem cards", "Số nhà/phòng/lấp đầy/doanh thu…", "Major", "Functional", "Landlord")
add("SH-D", M, "Dashboard tenant cards", "Thẻ tenant", "P1", "Login tenant", "—",
    "1. Xem", "Phòng đang thuê; unpaid; appointments…", "Major", "Functional", "Tenant")
add("SH-D", M, "Dashboard manager cards", "Thẻ manager", "P1", "Login manager", "—",
    "1. Xem", "ManagedHouses; pending maintenance…", "Major", "Functional", "Manager")
add("SH-D", M, "Occupancy chart", "Biểu đồ lấp đầy", "P2", "Landlord/Admin có room data", "—",
    "1. Xem chart panel", "Occupied/Available/Maintenance", "Minor", "Functional", "Landlord", "No")
add("SH-CAL", M, "Calendar events", "Lịch sự kiện theo role", "P1", "Có appointment/contract", "—",
    "1. Mở Lịch", "Events trong khoảng ngày", "Major", "Functional", "All")
add("SH-REP", M, "Report open", "Báo cáo Admin/Landlord", "P1", "Login role 1 hoặc 2", "—",
    "1. Mở Báo cáo", "Cards + sections; nút CSV/PDF không đè", "Major", "Functional", "Landlord")
add("SH-REP", M, "Export CSV", "Xuất CSV", "P2", "Report", "Path hợp lệ",
    "1. Xuất Excel CSV", "File tạo; mở được", "Minor", "Functional", "Admin", "No")
add("SH-REP", M, "Export HTML/PDF", "Xuất HTML", "P2", "Report", "—",
    "1. Xuất PDF/HTML", "File/browser OK", "Minor", "Functional", "Admin", "No")
add("SH-REP", M, "Tenant no report", "Tenant không báo cáo", "P1", "Login tenant", "—",
    "1. Menu", "Không có Reports", "Major", "Authorization", "Tenant")

# =============================================================================
# DATABASE / CONSTRAINT / TRANSACTION
# =============================================================================
M = "Database & Integrity"
add("DB", M, "UQ Assignments", "Unique HouseID+ManagerID", "P0", "SQL", "Insert trùng pair Active",
    "1. Insert trùng qua UI/SQL", "UQ_Assignments_House_Manager ngăn hoặc app reactivate", "Critical", "Database", "All", "Yes")
add("DB", M, "UQ Room number", "Unique House+RoomNumber", "P0", "SQL/UI", "Trùng số phòng",
    "1. Insert", "Constraint fail", "Critical", "Database", "Landlord", "Yes")
add("DB", M, "UQ Favorite", "Unique User+Room", "P1", "SQL", "Duplicate favorite",
    "1. Insert 2 lần", "UQ fail", "Major", "Database", "Tenant", "Yes")
add("DB", M, "FK Restrict user delete", "Xóa User còn House", "P1", "SQL", "Delete Owner còn nhà",
    "1. DELETE Users", "FK Restrict fail (không orphan)", "Major", "Database", "Admin", "Yes")
add("DB", M, "CK Room status", "Status phòng invalid", "P1", "SQL", "Status='Foo'",
    "1. UPDATE", "CHECK fail", "Major", "Database", "All", "Yes")
add("DB", M, "CK Contract status", "Status HĐ invalid", "P1", "SQL", "Status='Open'",
    "1. UPDATE", "CHECK fail (Draft|Active|Expired|Terminated)", "Major", "Database", "All", "Yes")
add("DB", M, "CK Meter New>=Old", "NewElectric < OldElectric", "P0", "SQL", "Violate check",
    "1. INSERT reading invalid", "CK_MeterReadings_Electric fail", "Critical", "Database", "Manager", "Yes")
add("DB", M, "Nullable TenantID draft", "Draft TenantID NULL", "P0", "UI tạo draft", "—",
    "1. Tạo nháp\n2. SELECT", "TenantID IS NULL; Status Draft", "Critical", "Database", "Landlord", "Yes")
add("DB", M, "Invoice transaction", "Generate invoice rollback khi lỗi sau reading", "P1",
    "Force lỗi (dev)", "—",
    "1. Quan sát transaction Begin/Commit", "Không còn reading mồ côi nếu fail giữa chừng", "Critical", "Transaction", "Manager", "Yes")
add("DB", M, "Contract create transaction", "CreateContract rollback", "P1", "Force fail", "—",
    "1. Transaction", "Không HĐ + Occupied lệch", "Critical", "Transaction", "Landlord", "Yes")
add("DB", M, "Pending columns exist", "Schema updater thêm cột pending", "P0", "App start lần đầu DB cũ", "—",
    "1. Start app\n2. sp_columns Contracts", "Có Pending* và Previous* PriceEffectiveDate", "Critical", "Regression", "All", "Yes")
add("DB", M, "Chat tables exist", "ChatConversations tạo nếu thiếu", "P1", "DB không có chat", "—",
    "1. Start app", "Bảng chat tồn tại", "Major", "Regression", "All", "Yes")

# =============================================================================
# STATUS TRANSITIONS
# =============================================================================
M = "Status Transitions"
add("ST", M, "Appointment flow", "Pending→Accepted→Completed", "P0", "Appointment", "—",
    "1. Accept\n2. Complete", "Chuyển đúng; mỗi bước có notify", "Critical", "Workflow", "Landlord")
add("ST", M, "Appointment reject", "Pending→Rejected", "P0", "Pending", "—",
    "1. Reject", "Rejected; không Complete sau", "Critical", "Workflow", "Landlord")
add("ST", M, "Contract draft to active", "Draft→Active qua AssignTenant", "P0", "Draft", "—",
    "1. Gán khách", "Active + Occupied", "Critical", "Workflow", "Landlord")
add("ST", M, "Contract terminate", "Active→Terminated", "P1", "Active", "—",
    "1. Terminate", "Terminated; Room Available", "Major", "Workflow", "Landlord")
add("ST", M, "Maintenance flow", "Pending→Processing→Completed", "P0", "Request", "—",
    "1. Process\n2. Complete", "Đúng chuỗi; CompletedDate", "Critical", "Workflow", "Manager")
add("ST", M, "Invoice Unpaid to Paid", "Unpaid→Paid", "P0", "Unpaid", "—",
    "1. Pay", "Paid + Payment", "Critical", "Workflow", "Tenant")
add("ST", M, "Post Pending to Approved", "Pending→Approved", "P0", "Pending post", "—",
    "1. Admin approve", "Approved; hiện search", "Critical", "Workflow", "Admin")
add("ST", M, "Assignment Active Inactive", "Active↔Inactive", "P0", "Assignment", "—",
    "1. Ngưng\n2. Gán lại", "Inactive rồi Active", "Critical", "Workflow", "Landlord")
add("ST", M, "User Active Inactive", "Toggle user", "P0", "User", "—",
    "1. Toggle\n2. Login", "Inactive chặn login", "Critical", "Workflow", "Admin")
add("ST", M, "Invalid jump maintenance", "Pending→Completed bỏ Processing (nếu UI cho)", "P2", "Pending", "—",
    "1. Thử Complete trực tiếp", "Theo rule: cho phép hoặc chặn — ghi kết quả thực tế", "Minor", "Exploratory", "Manager", "No")

# =============================================================================
# SECURITY / USABILITY / PERF / SMOKE / REGRESSION PACKS
# =============================================================================
M = "Security & Abuse"
add("SEC", M, "Password storage", "Password không plain trong DB sau seed", "P0", "App đã chạy DataSeeder", "Users.Password",
    "1. SELECT Password", "Hash $2a$/$2b$… không phải plain 123456", "Critical", "Security", "All", "Yes")
add("SEC", M, "IDOR contract", "Tenant đoán ContractID người khác", "P0", "2 tenant", "ContractID khác",
    "1. UI chỉ list theo TenantID", "Không thấy/sửa HĐ người khác", "Critical", "Security", "Tenant")
add("SEC", M, "IDOR invoice pay", "Pay invoice không thuộc mình", "P0", "Invoice tenant khác", "—",
    "1. Không lộ trên UI; service nên check", "Không thanh toán được HĐ người khác", "Critical", "Security", "Tenant", "Yes")
add("SEC", M, "SQL injection search", "Ô tìm phòng", "P1", "TenantHome", "'; DROP TABLE",
    "1. Tìm", "An toàn", "Critical", "Security", "Tenant")
add("SEC", M, "Path traversal upload", "Tên file ../", "P1", "Upload ảnh", "filename ../etc/passwd",
    "1. Upload", "Sanitize path; không ghi ngoài uploads", "Critical", "Security", "Landlord", "No")

M = "Usability & UI"
add("UX", M, "MainForm topbar", "User info / logout không đè title", "P1", "Tên dài", "Resize",
    "1. Resize MainForm", "Ellipsis; logout bấm được", "Major", "Usability", "All", "No")
add("UX", M, "Toast success", "Toast sau thao tác", "P2", "Gán manager/tạo HĐ", "—",
    "1. Thao tác thành công", "ToastNotifier hiện", "Minor", "Usability", "Landlord", "No")
add("UX", M, "Loading panel", "Dashboard loading", "P2", "Dashboard", "—",
    "1. Mở Dashboard", "Loading rồi ẩn", "Minor", "Usability", "All", "No")
add("UX", M, "Empty state meter", "Empty state có hướng dẫn", "P1", "Manager chưa gán", "—",
    "1. Mở Meter", "Hint phân công / HĐ Active", "Major", "Usability", "Manager", "No")
add("UX", M, "Dialog footer modal", "Save/Cancel không cắt khi resize modal", "P1", "Room/House/User modal", "Thu nhỏ",
    "1. Resize dialog", "Footer Dock Bottom còn nút", "Major", "Usability", "Admin", "No")
add("UX", M, "Form open regression", "Mọi menu mở được sau UI refactor", "P0", "Login từng role", "Mọi tag menu",
    "1. Click lần lượt mọi menu", "Không 'Không mở được màn hình'", "Critical", "Regression", "All", "No")

M = "Performance & Concurrency"
add("PERF", M, "Search many posts", "Tìm với nhiều tin", "P2", "Seed/ thêm 100 posts", "—",
    "1. Tìm kiếm", "Trả kết quả < 5s (ước lượng máy local)", "Minor", "Performance", "Tenant", "No")
add("PERF", M, "Grid large contracts", "Nhiều HĐ", "P2", "100 HĐ", "—",
    "1. Mở list HĐ", "UI vẫn scroll được", "Minor", "Performance", "Landlord", "No")
add("CONC", M, "Two landlords assign same", "Không conflict cross-owner", "P1", "2 session", "—",
    "1. Mỗi người gán nhà mình", "OK độc lập", "Major", "Concurrency", "Landlord", "No")
add("CONC", M, "DbContext concurrent UI", "Contract form đổi house nhanh", "P1", "LandlordContractForm", "Đổi combo liên tục",
    "1. Đổi nhà/phòng nhanh", "Không crash DbContext concurrency (scope)", "Critical", "Concurrency", "Landlord", "No")

M = "Smoke Suite"
for role, steps in [
    ("Admin", "Login → Users → Posts → Dashboard → Logout"),
    ("Landlord", "Login → Houses → Rooms → Contract → Appointment → Assignment → Logout"),
    ("Tenant", "Login → Search → Favorite → Contract → Invoice → Logout"),
    ("Manager", "Login → Meter → Maintenance → Notifications → Logout"),
]:
    add("SMOKE", M, f"Smoke {role}", f"Luồng smoke {role}", "P0", f"DB seed; account {role}", "Demo accounts",
        f"1. {steps}", "Mọi bước không lỗi blocker", "Critical", "Smoke", role, "No")

M = "Regression Hotspots"
add("REG", M, "Schema pending columns", "Query Contracts sau updater", "P0", "App start", "—",
    "1. Mở Meter/Contract", "Không Invalid column name Pending*", "Critical", "Regression", "All")
add("REG", M, "Page header label", "GetPageHeaderTitle", "P0", "Dashboard", "—",
    "1. Mở Dashboard", "Không Sequence contains no elements", "Critical", "Regression", "All")
add("REG", M, "Appointment notify fields", "UpdatedDate/IsRead trên notify", "P0", "Accept appointment", "—",
    "1. Accept\n2. SELECT Notifications", "IsRead=0; UpdatedDate hợp lệ; Tenant nhận", "Critical", "Regression", "Landlord")
add("REG", M, "Bulk draft codes unique", "ContractCode không trùng giây", "P1", "Bulk nhiều phòng", "—",
    "1. Bulk\n2. Check codes", "Unique nhờ RoomID+seq", "Major", "Regression", "Landlord", "Yes")
add("REG", M, "Manager deactivate lose access", "Ngưng gán → không còn HĐ", "P0", "Đã gán", "—",
    "1. Ngưng\n2. Manager refresh meter", "Empty / không còn HĐ nhà đó", "Critical", "Regression", "Manager")

# =============================================================================
# EDGE / EXPLORATORY / LOGGING / EXCEPTION
# =============================================================================
M = "Edge Cases & Logging"
add("EDGE", M, "Unicode Vietnamese", "Tên nhà tiếng Việt", "P1", "Create house", "Nhà trọ Nguyễn Văn Nam",
    "1. Lưu\n2. Xem lại", "Lưu đúng Unicode (DB NVARCHAR)", "Major", "Functional", "Landlord")
add("EDGE", M, "Null phone email", "User Phone/Email null", "P2", "Create user", "Phone/Email trống nếu cho phép",
    "1. Lưu", "OK nếu nullable; unique email skip null", "Minor", "Boundary", "Admin")
add("EDGE", M, "Move-out before month", "MoveOut khiến 0 ngày ở", "P1", "HĐ có MoveOut sớm", "Billing month ngoài ở",
    "1. Generate invoice", "BadRequest không có ngày ở trong tháng", "Major", "Business Rule", "Manager")
add("EDGE", M, "Electric price zero", "ElectricPrice=0", "P2", "Create contract", "Electric=0 Water>0",
    "1. Tạo HĐ\n2. Invoice", "Cho phép; cost điện=0", "Minor", "Boundary", "Landlord")
add("EDGE", M, "Activity log on login", "Ghi log đăng nhập", "P1", "Login", "—",
    "1. Login\n2. Admin ActivityLog", "Có Action đăng nhập", "Major", "Logging", "Admin")
add("EDGE", M, "Unhandled exception UI", "Lỗi DB khi SQL tắt", "P1", "Stop SQL service", "—",
    "1. Thao tác load list", "AppDialog lỗi; không crash process (ideal)", "Major", "Exception", "All", "No")
add("EDGE", M, "Refresh during load", "Click Làm mới liên tục", "P2", "List form", "—",
    "1. Spam refresh", "Không crash; data cuối cùng đúng", "Minor", "Concurrency", "All", "No")
add("EDGE", M, "Exploratory landlord day", "Khám phá 30 phút landlord", "P2", "Full seed", "—",
    "1. Tự do tạo nhà/phòng/HĐ/hẹn/gán manager", "Ghi bug phát hiện; không blocker unlogged", "Minor", "Exploratory", "Landlord", "No")
add("EDGE", M, "Exploratory tenant day", "Khám phá tenant", "P2", "Full seed", "—",
    "1. Search→book→pay→review", "Ghi nhận UX issues", "Minor", "Exploratory", "Tenant", "No")

# ---------------------------------------------------------------------------
# PART EXTRA — mở rộng đạt 400+ TC (equivalence / negative / role / DB)
# ---------------------------------------------------------------------------
M = "Auth Extended"
add("AUTH", M, "Show password checkbox", "Hiện/ẩn mật khẩu Login", "P2", "LoginForm", "—",
    "1. Tick hiện mật khẩu\n2. Bỏ tick", "Password mask/unmask đúng", "Minor", "Usability", "All", "No")
add("AUTH", M, "Login after logout other role", "Đổi role trong cùng process", "P0", "Login Admin rồi Logout", "tenant/123456",
    "1. Logout\n2. Login tenant", "Menu chuyển đúng Tenant; không sót quyền Admin", "Critical", "Regression", "Tenant")
add("AUTH", M, "Whitespace username", "Username có khoảng trắng đầu cuối", "P2", "Login", "  namlandlord  ",
    "1. Login", "Trim hoặc fail — ghi nhận", "Minor", "Boundary", "Landlord")
add("AUTH", M, "Case sensitivity username", "ADMIN vs admin", "P1", "Login", "ADMIN / admin123",
    "1. Login", "Theo DB collation/so khớp chính xác Username", "Major", "Negative", "Admin")
add("AUTH", M, "Concurrent two app instances", "2 instance cùng user", "P2", "Mở 2 app", "cùng account",
    "1. Login cả hai", "Cả hai vào được (desktop); ghi nhận conflict nếu có", "Minor", "Concurrency", "All", "No")

M = "Admin User Extended"
add("ADM-U", M, "Create Landlord", "Tạo role Landlord", "P0", "Admin", "Role=Landlord; username mới",
    "1. Tạo\n2. Login user mới", "Menu Landlord đúng", "Critical", "Functional", "Admin")
add("ADM-U", M, "Create Tenant", "Tạo role Tenant", "P0", "Admin", "Role=Tenant",
    "1. Tạo\n2. Login", "Menu Tenant", "Critical", "Functional", "Admin")
add("ADM-U", M, "Change role of existing", "Đổi RoleID user", "P1", "User Tenant", "Đổi thành Manager",
    "1. Update Role\n2. Login lại", "Menu theo role mới", "Major", "Functional", "Admin")
add("ADM-U", M, "Lock self admin", "Admin tự khóa chính mình", "P1", "Login admin", "Toggle Inactive chính mình",
    "1. Khóa\n2. Logout\n2. Login lại", "Không login được — rủi ro lockout (ghi nhận)", "Major", "Security", "Admin", "No")
add("ADM-U", M, "Search no match", "Keyword không khớp", "P2", "List users", "xyznotfound",
    "1. Tìm", "Grid rỗng", "Minor", "Functional", "Admin")
add("ADM-U", M, "Edit username disabled", "Username không sửa khi edit", "P1", "Edit modal", "—",
    "1. Mở sửa user", "txtUsername disabled", "Major", "UI Validation", "Admin", "No")

M = "Landlord House Room Extended"
add("LL-H", M, "List empty", "Landlord chưa có nhà", "P1", "Landlord mới", "—",
    "1. Mở Nhà của tôi", "Grid rỗng / empty hint", "Major", "Functional", "Landlord")
add("LL-H", M, "Special char house name", "Tên nhà có dấu & ký tự", "P2", "Create", "Nhà A&B <Test>",
    "1. Lưu", "Lưu đúng; không XSS", "Minor", "Security", "Landlord")
add("LL-R", M, "Bedroom bathroom zero", "Bedroom=0 Bathroom=0 hợp lệ", "P2", "Create room", "0/0",
    "1. Lưu", "OK theo CHECK >=0", "Minor", "Boundary", "Landlord")
add("LL-R", M, "Floor null", "Floor để trống", "P2", "Create", "Floor null",
    "1. Lưu", "OK nullable", "Minor", "Boundary", "Landlord")
add("LL-R", M, "Furniture long", "Furniture 500+ chars", "P2", "Create", "chuỗi dài",
    "1. Lưu", "MaxLength 500 hoặc lỗi", "Minor", "Boundary", "Landlord")
add("LL-R", M, "Same room number different houses", "101 ở 2 nhà", "P0", "2 nhà", "RoomNumber=101 mỗi nhà",
    "1. Tạo ở nhà A và B", "Cả hai OK (UQ theo HouseID)", "Critical", "Business Rule", "Landlord")
add("LL-R", M, "Update price occupied room", "Đổi Price phòng đang Occupied", "P1", "Occupied", "Price mới",
    "1. Sửa phòng", "Cho phép đổi snapshot phòng; HĐ giữ MonthlyRent cũ trừ khi sửa HĐ", "Major", "Business Rule", "Landlord")

M = "Contract Extended"
add("LL-C", M, "Create without house", "Chưa chọn nhà", "P1", "Form HĐ", "cboHouse trống",
    "1. Lưu", "Warning chọn phòng/nhà", "Major", "UI Validation", "Landlord")
add("LL-C", M, "Electric water decimals", "Giá điện thập phân", "P2", "Form", "Electric=3500.5",
    "1. Lưu HĐ\n2. Invoice", "Tính đúng decimal", "Minor", "Boundary", "Landlord")
add("LL-C", M, "Deposit equal rent", "Cọc = 1 tháng", "P2", "Form", "Deposit=MonthlyRent",
    "1. Lưu", "OK", "Minor", "Equivalence", "Landlord")
add("LL-C", M, "Assign inactive tenant", "Gán tenant Inactive", "P0", "Tenant bị khóa", "TenantID Inactive",
    "1. AssignTenant", "BadRequest không hoạt động", "Critical", "Negative", "Landlord")
add("LL-C", M, "Tenant confirm other contract", "Confirm Pending HĐ người khác", "P0", "Pending của tenant A; login B", "—",
    "1. B không thấy đề xuất A", "Không confirm được HĐ người khác", "Critical", "Security", "Tenant")
add("LL-C", M, "Cancel pending by other landlord", "Landlord khác hủy Pending", "P0", "Pending HĐ landlord A", "Login B",
    "1. B không thấy HĐ A", "Không hủy được", "Critical", "Security", "Landlord")
add("LL-C", M, "Extend end before start", "NewEnd < Start", "P1", "Extend", "NewEnd invalid",
    "1. Gia hạn", "BadRequest", "Major", "Negative", "Landlord")
add("LL-C", M, "Bulk with shared rent", "Bulk dùng Rent form >0", "P1", "Nhiều phòng giá khác nhau", "Rent shared=3000000",
    "1. Bulk", "Mọi Draft MonthlyRent=3000000", "Major", "Functional", "Landlord")
add("LL-C", M, "Print without tenant", "In HĐ Draft", "P2", "Draft", "—",
    "1. In/PDF", "In được với '(Chưa có khách)'", "Minor", "Functional", "Landlord", "No")

M = "Invoice Extended"
add("MG-M", M, "OtherFee large", "Phụ phí lớn", "P2", "Generate", "OtherFee=10000000",
    "1. Tạo HĐ", "Total = rent+e+w+fee", "Minor", "Boundary", "Manager")
add("MG-M", M, "OtherFee negative", "Phụ phí âm", "P1", "Form", "OtherFee=-1",
    "1. Tạo", "Validation không cho", "Major", "Negative", "Manager")
add("MG-M", M, "Refresh after generate", "List sau tạo HĐ", "P1", "Vừa tạo invoice", "—",
    "1. Tạo xong", "Form reset selection; vẫn chọn HĐ được", "Major", "Functional", "Manager", "No")
add("TN-I", M, "Unpaid only pay button", "Chỉ Unpaid thanh toán", "P0", "Mix Paid/Unpaid", "—",
    "1. Thử pay Paid", "Không cho", "Critical", "Business Rule", "Tenant")
add("TN-I", M, "Invoice of terminated contract", "HĐ đã Terminated còn Unpaid", "P1", "Terminate sau khi có Unpaid", "—",
    "1. Tenant vẫn thấy Unpaid?\n2. Có thể pay?", "Theo rule: vẫn thu nợ hoặc khóa — ghi nhận", "Major", "Business Rule", "Tenant")

M = "Appointment Extended"
add("LL-AP", M, "Accept already accepted", "Nhận lại lịch Accepted", "P1", "Accepted", "—",
    "1. Nhận lại", "OK idempotent hoặc thông báo", "Minor", "Negative", "Landlord")
add("LL-AP", M, "Filter All status", "Status=All", "P1", "Nhiều status", "All",
    "1. Lọc", "Hiện mọi status trong date range", "Major", "Functional", "Landlord")
add("TN-AP", M, "Book without login", "Không áp dụng — bắt buộc login", "P0", "Logout", "—",
    "1. Không vào được TenantHome", "Bắt login", "Critical", "Security", "Tenant")
add("TN-AP", M, "Book rejected room again", "Đặt lại sau Reject", "P1", "Lịch Rejected", "Room cũ",
    "1. Đặt lịch mới", "Tạo Appointment Pending mới được", "Major", "Functional", "Tenant")

M = "Post Extended"
add("LL-PO", M, "Create post occupied room", "Đăng tin phòng Occupied", "P1", "Occupied", "—",
    "1. Đăng tin", "Cho phép hoặc chặn — ghi nhận nghiệp vụ", "Major", "Business Rule", "Landlord")
add("ADM-P", M, "Reject then approve", "Rejected không duyệt lại?", "P2", "Rejected", "—",
    "1. Approve", "Theo implement", "Minor", "Workflow", "Admin")
add("TN-S", M, "Rejected post hidden", "Tin Rejected không search", "P0", "Post Rejected", "—",
    "1. Tenant tìm", "Không hiện", "Critical", "Business Rule", "Tenant")
add("TN-S", M, "Pending post hidden", "Tin Pending không search", "P0", "Post Pending", "—",
    "1. Tenant tìm", "Không hiện đến khi Approved", "Critical", "Business Rule", "Tenant")

M = "Maintenance Extended"
add("TN-M", M, "Description only spaces", "Mô tả khoảng trắng", "P2", "Form", "Description='   '",
    "1. Gửi", "Validation", "Minor", "UI Validation", "Tenant")
add("MG-MT", M, "Complete without accept", "Complete từ Pending", "P1", "Pending", "—",
    "1. Bấm Xong", "Cho phép hoặc yêu cầu Processing trước", "Major", "Workflow", "Manager")
add("MG-MT", M, "Refresh list", "Làm mới sau tenant tạo", "P1", "Manager form mở", "Tenant tạo request",
    "1. Làm mới", "Thấy request mới", "Major", "Functional", "Manager", "No")

M = "Chat Calendar Extended"
add("SH-C", M, "Long message", "Tin nhắn 4000 ký tự", "P2", "Chat", "Content max",
    "1. Gửi", "OK hoặc cắt theo nvarchar(4000)", "Minor", "Boundary", "Tenant")
add("SH-C", M, "Rapid send", "Gửi liên tục 10 tin", "P2", "Chat", "—",
    "1. Spam gửi", "Không mất tin; không crash", "Minor", "Concurrency", "Landlord", "No")
add("SH-CAL", M, "Empty range", "Không sự kiện", "P2", "Date range trống data", "—",
    "1. Mở lịch", "Empty / không crash", "Minor", "Functional", "All")

M = "Dashboard Report Extended"
add("SH-D", M, "Admin dashboard numbers", "Cards admin", "P1", "Login admin", "—",
    "1. Xem TotalUsers/Houses/Posts…", "Số >= seed tối thiểu", "Major", "Functional", "Admin")
add("SH-D", M, "Revenue chart empty", "Landlord chưa doanh thu", "P2", "Landlord mới", "—",
    "1. Dashboard", "Chart rỗng không crash", "Minor", "Functional", "Landlord")
add("SH-REP", M, "Header buttons resize", "CSV/PDF không đè title", "P1", "ReportForm thu nhỏ", "—",
    "1. Resize", "Nút vẫn dùng (CreatePageHeader)", "Major", "Usability", "Admin", "No")

M = "Database Extended"
add("DB", M, "FK appointment tenant", "TenantID không tồn tại", "P1", "SQL", "TenantID=99999",
    "1. INSERT Appointment", "FK fail", "Major", "Database", "All", "Yes")
add("DB", M, "CK invoice status", "Status invoice invalid", "P1", "SQL", "Status='Waiting'",
    "1. UPDATE", "CHECK fail", "Major", "Database", "All", "Yes")
add("DB", M, "CK payment method", "Method invalid nếu có check", "P2", "SQL", "Method='Crypto'",
    "1. INSERT", "Fail hoặc OK tùy CK", "Minor", "Database", "All", "Yes")
add("DB", M, "Cascade room images", "Xóa Room xóa images", "P1", "Room có images", "—",
    "1. Delete room", "RoomImages CASCADE xóa", "Major", "Database", "Landlord", "Yes")
add("DB", M, "Review one per contract", "2 review cùng ContractID", "P0", "SQL", "Duplicate Review",
    "1. INSERT 2", "UQ/PK fail", "Critical", "Database", "Tenant", "Yes")
add("DB", M, "Notification UpdatedDate", "MinValue không insert", "P0", "App tạo notify", "—",
    "1. Accept appointment\n2. SELECT UpdatedDate", "UpdatedDate gần Now (không 0001)", "Critical", "Regression", "All", "Yes")

M = "Equivalence Partitioning Pack"
# Password length partitions
for label, pwd, exp in [
    ("Password empty", "", "Fail"),
    ("Password 1 char", "a", "Fail hoặc theo policy"),
    ("Password 6 char demo", "123456", "Pass demo accounts"),
    ("Password 72+ bcrypt", "x"*80, "Hash OK hoặc limit"),
]:
    add("EQ", M, label, f"EP password: {label}", "P2", "Login/ChangePwd", pwd,
        "1. Thử password partition", exp, "Minor", "Equivalence", "All")

M = "Role Direct Access Pack"
for role in ["Admin", "Landlord", "Tenant", "Manager"]:
    add("PERM", M, f"{role} only own data", f"{role} data isolation smoke", "P0", f"Login {role}", "—",
        f"1. Duyệt mọi màn {role}\n2. Xác nhận không thấy dữ liệu role khác trên UI",
        "Không lộ dữ liệu trái owner/tenant/assignment", "Critical", "Security", role, "No")

M = "Error Handling Pack"
add("ERR", M, "SQL down on login", "SQL Express tắt", "P1", "Stop SQL", "—",
    "1. Login", "Message lỗi DB rõ; không freeze vô hạn", "Major", "Exception", "All", "No")
add("ERR", M, "SQL down mid-form", "Mất kết nối khi load grid", "P1", "Form đang mở; stop SQL", "—",
    "1. Làm mới", "ShowError", "Major", "Exception", "Landlord", "No")
add("ERR", M, "Disk full export", "Xuất CSV ổ đầy (giả lập)", "P2", "Report", "path read-only",
    "1. Export", "Lỗi file IO thân thiện", "Minor", "Exception", "Admin", "No")

M = "Data Validation Extra"
for field, data, exp in [
    ("Deposit negative", "Deposit=-1", "Không cho số âm"),
    ("Water negative", "WaterPrice=-1", "Validation"),
    ("StartDate null UI", "Clear date nếu được", "Không lưu thiếu ngày"),
    ("RoomNumber empty", "RoomNumber=''", "Required"),
    ("HouseName only spaces", "HouseName='   '", "Trim/validate"),
    ("Username with spaces", "Username='a b'", "Reject hoặc trim rule"),
    ("Email invalid format", "Email=not-an-email", "Validate format nếu có"),
    ("Phone letters", "Phone=abcdef", "Validate hoặc cho phép lưu"),
]:
    add("VAL", M, field, f"Validate {field}", "P2", "Form liên quan", data,
        f"1. Nhập {data}\n2. Lưu", exp, "Minor", "UI Validation", "All")

M = "Permission Matrix Extra"
for role, action, exp in [
    ("Tenant", "Mở UI tạo nhà", "Không có menu"),
    ("Manager", "Duyệt post", "Không có menu Admin"),
    ("Landlord", "Ghi chỉ số meter", "Không có menu Manager"),
    ("Admin", "Gán manager qua LandlordAssignment", "Không có menu Landlord Assignment (đã chuyển landlord)"),
    ("Tenant", "Accept appointment", "Không có UI accept"),
    ("Manager", "Pay invoice", "Không có TenantInvoice"),
]:
    add("PERM", M, action, f"{role} không {action}", "P0", f"Login {role}", "—",
        f"1. Kiểm tra menu/UI không cho {action}", exp, "Critical", "Authorization", role)

# Notification flow matrix
M = "Notification Flows"
flows = [
    ("Book appointment", "Tenant đặt lịch", "Landlord nhận 'Có lịch hẹn…'", "Tenant", "Landlord"),
    ("Accept appointment", "Landlord Nhận", "Tenant nhận xác nhận", "Landlord", "Tenant"),
    ("Reject appointment", "Landlord Từ chối", "Tenant nhận từ chối", "Landlord", "Tenant"),
    ("Create contract with tenant", "Landlord tạo Active", "Tenant nhận HĐ mới", "Landlord", "Tenant"),
    ("Assign tenant", "Gán khách Draft", "Tenant nhận notify", "Landlord", "Tenant"),
    ("Pending contract edit", "Sửa Active", "Tenant nhận đề xuất", "Landlord", "Tenant"),
    ("Generate invoice", "Manager tạo HĐ tháng", "Tenant nhận hóa đơn", "Manager", "Tenant"),
    ("Pay invoice", "Tenant thanh toán", "Có thể notify landlord (nếu implement) — ghi giả định", "Tenant", "Landlord"),
    ("Assign manager", "Landlord gán", "Manager nhận được gán nhà", "Landlord", "Manager"),
    ("Maintenance status", "Manager cập nhật", "Tenant nhận phản hồi BT", "Manager", "Tenant"),
]
for feat, step, exp, actor, receiver in flows:
    add("NF", M, feat, f"Notify: {feat}", "P0", "Data đủ để thao tác", step,
        f"1. {actor}: {step}\n2. Login {receiver}\n3. Mở Thông báo",
        exp, "Critical", "Functional", actor)

# ---------------------------------------------------------------------------
# PART EXTRA 2 — đẩy ≥450 TC: Register, Profile, Favorite, Review, Backup,
# Detail forms, Status matrix, Usability, Performance
# ---------------------------------------------------------------------------
M = "Registration"
add("REGU", M, "Open register from login", "Mở form đăng ký", "P0", "LoginForm", "—",
    "1. Click Đăng ký / link Register", "Mở RegisterForm", "Critical", "Smoke", "All", "No")
add("REGU", M, "Register success", "Đăng ký user mới", "P0", "Username chưa tồn tại", "user mới + password đủ",
    "1. Điền form\n2. Đăng ký\n3. Login", "Tạo user Active; login được", "Critical", "Functional", "All")
add("REGU", M, "Duplicate username", "Username trùng", "P0", "Username đã có", "admin",
    "1. Đăng ký với username trùng", "Lỗi unique; không tạo", "Critical", "Negative", "All")
add("REGU", M, "Duplicate email", "Email trùng", "P1", "Email đã dùng", "email trùng",
    "1. Đăng ký", "Lỗi unique email (nếu bắt buộc)", "Major", "Negative", "All")
add("REGU", M, "Password mismatch confirm", "Confirm password khác", "P1", "RegisterForm", "pwd ≠ confirm",
    "1. Submit", "Validation không tạo user", "Major", "UI Validation", "All")
add("REGU", M, "Empty required fields", "Thiếu FullName/Username", "P1", "RegisterForm", "để trống",
    "1. Submit", "Không tạo; hiện required", "Major", "UI Validation", "All")
add("REGU", M, "Default role tenant", "Role mặc định khi đăng ký", "P0", "Register", "user mới",
    "1. Đăng ký\n2. Login\n3. Kiểm tra menu", "Role = Tenant (giả định) hoặc theo UI Role combo", "Critical", "Business Rule", "All")
add("REGU", M, "Cannot self-register Admin", "Không tự chọn Admin", "P0", "RegisterForm", "—",
    "1. Kiểm tra combo Role", "Không cho chọn Admin (hoặc Admin chỉ seed)", "Critical", "Security", "All", "No")

M = "Profile & Password"
add("PROF", M, "View profile", "Xem hồ sơ", "P0", "Đã login", "—",
    "1. Mở Hồ sơ / Profile", "Hiện Username, FullName, Email, Phone, Role", "Critical", "Functional", "All")
add("PROF", M, "Update full name", "Sửa họ tên", "P1", "ProfileForm", "FullName mới",
    "1. Sửa\n2. Lưu\n3. Mở lại", "Lưu thành công", "Major", "Functional", "Landlord")
add("PROF", M, "Update email unique", "Đổi email trùng user khác", "P1", "2 user", "Email của user khác",
    "1. Lưu", "Lỗi unique", "Major", "Negative", "Tenant")
add("PROF", M, "Change password success", "Đổi mật khẩu đúng", "P0", "Biết mật khẩu cũ", "old đúng; new hợp lệ",
    "1. Đổi MK\n2. Logout\n3. Login new", "Login bằng mật khẩu mới OK", "Critical", "Functional", "All")
add("PROF", M, "Change password wrong old", "Sai mật khẩu cũ", "P0", "Profile", "old sai",
    "1. Đổi MK", "Từ chối; MK không đổi", "Critical", "Negative", "All")
add("PROF", M, "Change password empty new", "MK mới trống", "P1", "Profile", "new=''",
    "1. Submit", "Validation", "Major", "UI Validation", "All")
add("PROF", M, "Username readonly", "Username không đổi trên Profile", "P1", "Profile", "—",
    "1. Kiểm tra field Username", "Readonly / disabled", "Major", "UI Validation", "All", "No")

M = "Favorite & Review"
add("TN-F", M, "Add favorite from detail", "Thêm yêu thích", "P0", "Tenant; post Approved", "Room có post",
    "1. RoomDetail\n2. Yêu thích\n3. Mở Favorites", "Xuất hiện trong list", "Critical", "Functional", "Tenant")
add("TN-F", M, "Remove favorite", "Bỏ yêu thích", "P1", "Đã favorite", "—",
    "1. Bỏ yêu thích\n2. Refresh list", "Không còn trong Favorites", "Major", "Functional", "Tenant")
add("TN-F", M, "Duplicate favorite", "Favorite 2 lần cùng room", "P1", "Đã favorite", "—",
    "1. Add lại", "Idempotent / không duplicate row", "Major", "Negative", "Tenant")
add("TN-F", M, "Favorite deleted post", "Post bị reject sau khi favorite", "P2", "Đã favorite; Admin reject", "—",
    "1. Mở Favorites", "Ẩn hoặc đánh dấu không khả dụng", "Minor", "Business Rule", "Tenant")
add("TN-RV", M, "Create review after active", "Đánh giá sau HĐ Active", "P0", "Tenant có Active contract", "Rating 1-5 + comment",
    "1. TenantReview\n2. Gửi", "Tạo Review; landlord thấy", "Critical", "Functional", "Tenant")
add("TN-RV", M, "Review without contract", "Đánh giá khi chưa thuê", "P0", "Tenant không HĐ", "—",
    "1. Thử tạo review", "Không cho / list trống", "Critical", "Negative", "Tenant")
add("TN-RV", M, "Duplicate review same contract", "2 review cùng HĐ", "P0", "Đã review", "—",
    "1. Review lần 2", "Chặn unique ContractID", "Critical", "Business Rule", "Tenant")
add("TN-RV", M, "Rating boundary", "Rating 0 và 6", "P1", "Form review", "0 / 6",
    "1. Submit", "Chỉ nhận 1-5", "Major", "Boundary", "Tenant")
add("TN-RV", M, "Comment XSS", "Comment chứa script", "P2", "Review", "<script>alert(1)</script>",
    "1. Lưu\n2. Landlord xem", "Plain text; không execute", "Minor", "Security", "Tenant")
add("ADM-RV", M, "Admin view reviews", "Admin quản lý đánh giá", "P1", "Có reviews", "—",
    "1. Mở ReviewManagement", "List reviews; filter/search nếu có", "Major", "Functional", "Admin")
add("ADM-RV", M, "Admin delete review", "Xóa review vi phạm", "P1", "Có review", "—",
    "1. Xóa\n2. Tenant/Landlord refresh", "Review biến mất", "Major", "Functional", "Admin")
add("LL-RV", M, "Landlord see own reviews", "Chủ xem review nhà mình", "P1", "Có review trên room/house", "—",
    "1. LandlordReviewForm", "Chỉ review liên quan nhà mình", "Major", "Authorization", "Landlord")

M = "Room & Invoice Detail UI"
add("TN-S", M, "Open room detail", "Chi tiết phòng từ search", "P0", "Có post Approved", "—",
    "1. Search\n2. Double-click / Xem", "RoomDetailForm: giá, mô tả, ảnh, book, favorite", "Critical", "Functional", "Tenant")
add("TN-S", M, "Room detail no images", "Phòng không ảnh", "P2", "Room 0 images", "—",
    "1. Mở detail", "Placeholder; không crash", "Minor", "Functional", "Tenant", "No")
add("TN-I", M, "Invoice detail open", "Chi tiết hóa đơn", "P0", "Có invoice", "—",
    "1. TenantInvoice\n2. Xem chi tiết", "InvoiceDetail: rent, e, w, fee, total, status", "Critical", "Functional", "Tenant")
add("TN-I", M, "Pay from detail", "Thanh toán từ detail", "P1", "Unpaid invoice", "—",
    "1. Pay trên detail", "Status→Paid; Payment record", "Major", "Functional", "Tenant")

M = "Admin Backup & Activity"
add("ADM-B", M, "Open backup form", "Mở Backup (nếu có)", "P1", "Login Admin", "—",
    "1. Menu Backup", "Mở BackupForm hoặc báo thiếu module (ghi Blocked)", "Major", "Smoke", "Admin", "No")
add("ADM-B", M, "Create backup file", "Tạo backup DB", "P1", "BackupForm available; SQL online", "path hợp lệ",
    "1. Backup\n2. Kiểm tra file", "File .bak/.sql tồn tại", "Major", "Functional", "Admin", "No")
add("ADM-B", M, "Backup path invalid", "Path không ghi được", "P2", "BackupForm", "path read-only",
    "1. Backup", "Lỗi thân thiện", "Minor", "Negative", "Admin", "No")
add("ADM-B", M, "Non-admin backup menu", "Landlord không thấy Backup", "P0", "Login Landlord", "—",
    "1. Kiểm tra menu", "Không có Backup", "Critical", "Authorization", "Landlord")
add("ADM-L", M, "Activity log filter date", "Lọc ActivityLog theo ngày", "P1", "Có logs", "From-To",
    "1. Filter\n2. Apply", "Chỉ log trong khoảng", "Major", "Functional", "Admin")
add("ADM-L", M, "Activity log search user", "Tìm theo username", "P1", "Có logs", "keyword",
    "1. Search", "Khớp Action/User", "Major", "Functional", "Admin")
add("ADM-L", M, "Activity log empty range", "Khoảng không có data", "P2", "—", "năm 2099",
    "1. Filter", "Grid rỗng", "Minor", "Functional", "Admin")

M = "Status Transition Matrix"
# Contract
for st_from, st_to, actor, ok in [
    ("Draft", "Active", "Landlord Assign+Tenant confirm hoặc Create Active", True),
    ("Draft", "Cancelled", "Landlord cancel", True),
    ("Pending", "Active", "Tenant confirm", True),
    ("Pending", "Cancelled", "Landlord/Tenant reject", True),
    ("Active", "Terminated", "Landlord terminate", True),
    ("Active", "Draft", "Illegal revert", False),
    ("Terminated", "Active", "Illegal reactivate", False),
    ("Cancelled", "Active", "Illegal", False),
]:
    add("ST", M, f"Contract {st_from}->{st_to}", f"HĐ {st_from} → {st_to}", "P0" if not ok or st_to in ("Active","Terminated") else "P1",
        f"HĐ Status={st_from}", actor,
        f"1. Thực hiện chuyển {st_from}→{st_to}",
        "Thành công; audit/notify nếu có" if ok else "Bị chặn / không có UI",
        "Critical" if not ok else "Major", "Workflow", "Landlord" if "Landlord" in actor or "Illegal" in actor else "Tenant")

# Appointment / Maintenance / Invoice / Post / Room
for entity, fr, to, role, ok in [
    ("Appointment", "Pending", "Accepted", "Landlord", True),
    ("Appointment", "Pending", "Rejected", "Landlord", True),
    ("Appointment", "Accepted", "Rejected", "Landlord", False),
    ("Appointment", "Rejected", "Accepted", "Landlord", False),
    ("Maintenance", "Pending", "Processing", "Manager", True),
    ("Maintenance", "Processing", "Completed", "Manager", True),
    ("Maintenance", "Completed", "Pending", "Manager", False),
    ("Invoice", "Unpaid", "Paid", "Tenant", True),
    ("Invoice", "Paid", "Unpaid", "Tenant", False),
    ("Post", "Pending", "Approved", "Admin", True),
    ("Post", "Pending", "Rejected", "Admin", True),
    ("Post", "Approved", "Pending", "Admin", False),
    ("Room", "Available", "Occupied", "System on Active contract", True),
    ("Room", "Occupied", "Available", "System on Terminate", True),
    ("Assignment", "Active", "Inactive", "Landlord deactivate", True),
    ("Assignment", "Inactive", "Active", "Landlord re-assign / activate", True),
]:
    add("ST", M, f"{entity} {fr}->{to}", f"{entity} status {fr}→{to}", "P1",
        f"{entity}={fr}", f"Actor={role}",
        f"1. Thực hiện transition",
        "OK cập nhật DB+UI" if ok else "Không cho phép",
        "Major", "Workflow", role)

M = "Usability & Accessibility"
add("UX", M, "Tab order login", "Tab thứ tự Login", "P2", "LoginForm", "—",
    "1. Tab qua controls", "Username→Password→Login hợp lý", "Minor", "Usability", "All", "No")
add("UX", M, "Enter to login", "Enter submit login", "P2", "LoginForm", "credentials đúng",
    "1. Focus password\n2. Enter", "Đăng nhập", "Minor", "Usability", "All", "No")
add("UX", M, "Confirm dialog cancel", "Hủy dialog xác nhận", "P1", "Thao tác xóa/terminate", "—",
    "1. Bấm thao tác\n2. Cancel dialog", "Không thay đổi data", "Major", "Usability", "Landlord", "No")
add("UX", M, "Empty state hints", "Empty state có hướng dẫn", "P2", "List trống", "—",
    "1. Mở form không data", "Hint rõ ràng không chỉ grid trống", "Minor", "Usability", "Manager", "No")
add("UX", M, "DPI scaling", "UI ở 125% scaling", "P2", "Windows 125%", "—",
    "1. Mở các form chính", "Không cắt nút quan trọng (Assignment/Report)", "Major", "Usability", "All", "No")
add("UX", M, "Vietnamese labels", "Nhãn tiếng Việt nhất quán", "P2", "Mọi menu", "—",
    "1. Duyệt menu", "Không lẫn EN/VN lung tung trên cùng toolbar", "Minor", "Usability", "All", "No")

M = "Performance Basic"
add("PERF", M, "Login under 3s", "Login < 3s local", "P1", "SQL local", "demo account",
    "1. Đo thời gian login→MainForm", "<3s trên máy dev", "Major", "Performance", "All", "No")
add("PERF", M, "Search posts 100+", "Search với nhiều post", "P2", "Seed ≥100 posts nếu có", "keyword",
    "1. Search", "UI không treo >5s", "Minor", "Performance", "Tenant", "No")
add("PERF", M, "Dashboard load", "Dashboard < 5s", "P1", "Login", "—",
    "1. Mở Dashboard", "Cards/chart hiện <5s", "Major", "Performance", "Landlord", "No")
add("PERF", M, "Generate invoice calc", "Tạo HĐ 1 phòng nhanh", "P1", "Meter sẵn", "—",
    "1. Generate", "<2s + message", "Major", "Performance", "Manager", "No")
add("PERF", M, "Notification unread badge", "Badge cập nhật", "P2", "Có unread", "—",
    "1. Nhận notify\n2. Xem badge", "Số unread tăng (nếu có badge)", "Minor", "Performance", "Tenant", "No")

M = "Landlord Post CRUD Extra"
add("LL-PO", M, "Edit own pending post", "Sửa tin Pending", "P1", "Post Pending của mình", "Title mới",
    "1. Sửa\n2. Lưu", "Cập nhật; vẫn Pending", "Major", "Functional", "Landlord")
add("LL-PO", M, "Edit approved post", "Sửa tin đã Approved", "P1", "Approved", "Content mới",
    "1. Sửa", "Cho phép hoặc đưa về Pending — ghi nhận", "Major", "Business Rule", "Landlord")
add("LL-PO", M, "Delete post", "Xóa tin của mình", "P1", "Có post", "—",
    "1. Xóa\n2. Confirm", "Xóa khỏi list; tenant không search thấy", "Major", "Functional", "Landlord")
add("LL-PO", M, "Cannot edit other landlord post", "Không sửa tin người khác", "P0", "2 landlord", "—",
    "1. Login B\n2. List posts", "Không thấy/sửa post A", "Critical", "Authorization", "Landlord")
add("LL-PO", M, "Title empty", "Title trống", "P1", "Create post", "Title=''",
    "1. Lưu", "Validation required", "Major", "UI Validation", "Landlord")
add("LL-PO", M, "Title very long", "Title 300 chars", "P2", "Create", "chuỗi dài",
    "1. Lưu", "MaxLength hoặc lỗi", "Minor", "Boundary", "Landlord")

M = "Manager Meter Extra Cases"
add("MG-M", M, "Prev reading display", "Hiện chỉ số tháng trước", "P0", "Đã có invoice tháng trước", "—",
    "1. Chọn HĐ + tháng mới", "PrevElectric/Water đúng", "Critical", "Functional", "Manager")
add("MG-M", M, "First month zero prev", "Tháng đầu prev=0", "P0", "HĐ mới chưa invoice", "—",
    "1. Chọn tháng đầu", "Prev=0 hoặc từ contract start reading", "Critical", "Business Rule", "Manager")
add("MG-M", M, "Current less than prev", "Chỉ số hiện < trước", "P0", "Prev=100", "Current=50",
    "1. Generate", "BadRequest không hợp lệ", "Critical", "Negative", "Manager")
add("MG-M", M, "Equal prev current", "Current = Prev", "P1", "Prev=100", "Current=100",
    "1. Generate", "OK; usage=0; cost điện/nước=0", "Major", "Boundary", "Manager")
add("MG-M", M, "Partial month proration", "MoveIn giữa tháng", "P1", "StartDate mid-month", "tháng đó",
    "1. Generate", "Rent prorate theo số ngày", "Major", "Business Rule", "Manager")
add("MG-M", M, "House not assigned", "Nhà chưa gán manager", "P0", "Manager; nhà khác landlord chưa gán", "—",
    "1. Mở Meter", "Không thấy HĐ nhà đó", "Critical", "Authorization", "Manager")
add("MG-M", M, "Double generate same month", "Tạo trùng tháng", "P0", "Đã có invoice tháng", "cùng tháng",
    "1. Generate lại", "Chặn duplicate month", "Critical", "Business Rule", "Manager")
add("MG-M", M, "Notes field", "Ghi chú hóa đơn", "P2", "Generate", "Notes dài",
    "1. Tạo", "Lưu Notes nếu có field", "Minor", "Functional", "Manager")

M = "Chat Security Extra"
add("SH-C", M, "Cannot open arbitrary conversation", "Không chat user trái quyền", "P0", "Tenant", "Landlord không liên quan",
    "1. Thử mở chat ID khác", "Không có UI / từ chối", "Critical", "Security", "Tenant")
add("SH-C", M, "Empty message", "Gửi tin trống", "P1", "Chat mở", "Content=''",
    "1. Gửi", "Không gửi / validation", "Major", "UI Validation", "Landlord")
add("SH-C", M, "SQL in message", "Tin nhắn SQL", "P1", "Chat", "'; DROP TABLE Messages;--",
    "1. Gửi\n2. Reload", "Lưu plain text; DB intact", "Major", "Security", "Tenant")
add("SH-C", M, "Mark read", "Đánh dấu đã đọc", "P2", "Có unread", "—",
    "1. Mở hội thoại", "IsRead cập nhật nếu có", "Minor", "Functional", "All")

M = "Calendar Extra"
add("SH-CAL", M, "See own appointments", "Lịch hiện appointment của mình", "P1", "Có Accepted/Pending", "—",
    "1. Mở Calendar", "Thấy sự kiện đúng role", "Major", "Functional", "Landlord")
add("SH-CAL", M, "Filter date range", "Đổi khoảng ngày", "P1", "CalendarForm", "From-To",
    "1. Đổi range\n2. Refresh", "Sự kiện cập nhật", "Major", "Functional", "Tenant")
add("SH-CAL", M, "Click event detail", "Click sự kiện", "P2", "Có event", "—",
    "1. Click", "Hiện chi tiết hoặc navigate", "Minor", "Usability", "Manager", "No")

M = "House Soft Rules"
add("LL-H", M, "Delete house with rooms", "Xóa nhà còn phòng", "P0", "Nhà có phòng", "—",
    "1. Xóa nhà", "Chặn hoặc cascade theo rule; không orphan sai", "Critical", "Business Rule", "Landlord")
add("LL-H", M, "Delete house with active contract", "Xóa nhà đang thuê", "P0", "Có Active HĐ", "—",
    "1. Xóa", "Chặn", "Critical", "Negative", "Landlord")
add("LL-H", M, "Update address", "Sửa địa chỉ nhà", "P1", "Có nhà", "Address mới",
    "1. Sửa\n2. Lưu", "Cập nhật; post/search phản ánh", "Major", "Functional", "Landlord")
add("LL-H", M, "Inactive house hide posts", "Ngưng nhà → tin ẩn?", "P1", "House Inactive + post Approved", "—",
    "1. Tenant search", "Theo rule: ẩn hoặc vẫn hiện — ghi nhận", "Major", "Business Rule", "Tenant")

M = "Double-Submit & Race Pack"
add("RACE", M, "Double click create house", "Double-click Lưu nhà", "P1", "House modal", "data hợp lệ",
    "1. Double-click Save nhanh", "Chỉ 1 record (disable button / guard)", "Major", "Concurrency", "Landlord", "No")
add("RACE", M, "Double click pay", "Double-click Thanh toán", "P0", "Unpaid", "—",
    "1. Double-click Pay", "1 Payment; Status Paid 1 lần", "Critical", "Concurrency", "Tenant", "No")
add("RACE", M, "Double accept appointment", "Double Nhận lịch", "P1", "Pending", "—",
    "1. Double Accept", "1 lần Accepted; 1 notify", "Major", "Concurrency", "Landlord", "No")
add("RACE", M, "Refresh after pay", "F5/Refresh sau pay", "P1", "Vừa Paid", "—",
    "1. Refresh list", "Vẫn Paid; không pay lại", "Major", "Regression", "Tenant", "No")
add("RACE", M, "Session after logout navigate", "Sau logout không dùng form cũ", "P0", "Mở form rồi Logout", "—",
    "1. Logout\n2. Nếu form cũ còn, thao tác", "Không thao tác được / session null", "Critical", "Security", "All", "No")

print(f"Total cases generated: {len(cases)}")

# =============================================================================
# WRITE EXCEL
# =============================================================================
wb = Workbook()

header_fill = PatternFill("solid", fgColor="1E3A5F")
header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
wrap = Alignment(wrap_text=True, vertical="top")
thin = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)
prio_fills = {
    "P0": PatternFill("solid", fgColor="FECACA"),
    "P1": PatternFill("solid", fgColor="FDE68A"),
    "P2": PatternFill("solid", fgColor="E5E7EB"),
}


def write_sheet(ws, rows):
    ws.append(COLS)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    for row in rows:
        ws.append([row.get(c, "") for c in COLS])
    for r in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=len(COLS)):
        for cell in r:
            cell.alignment = wrap
            cell.border = thin
            cell.font = Font(name="Calibri", size=10)
        # priority color
        prio_cell = r[4]
        if prio_cell.value in prio_fills:
            prio_cell.fill = prio_fills[prio_cell.value]
    widths = [14, 28, 22, 40, 8, 36, 32, 40, 40, 14, 10, 10, 14, 12, 12]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 30


# Master
ws0 = wb.active
ws0.title = "00_All_TestCases"
write_sheet(ws0, cases)

# By module prefix groups
groups = {}
for c in cases:
    key = c["Test Case ID"].split("-")[1]
    # better group by Module
    groups.setdefault(c["Module"], []).append(c)

# Limit sheet name 31 chars; Excel forbids: \ / ? * [ ]
used_names = {"00_All_TestCases"}
idx = 1
for module, rows in groups.items():
    safe = (
        module.replace("/", "-")
        .replace("\\", "-")
        .replace("?", "")
        .replace("*", "")
        .replace("[", "")
        .replace("]", "")
        .replace(":", "-")
    )
    name = f"{idx:02d}_{safe}"[:31]
    base = name
    n = 1
    while name in used_names:
        name = (base[:27] + f"_{n}")[:31]
        n += 1
    used_names.add(name)
    ws = wb.create_sheet(name)
    write_sheet(ws, rows)
    idx += 1

# Summary sheet
ws_s = wb.create_sheet("99_Summary", 0)
ws_s["A1"] = "RPMS Test Case Suite — Summary"
ws_s["A1"].font = Font(bold=True, size=14, color="1E3A5F")
ws_s["A3"] = "Total Test Cases"
ws_s["B3"] = len(cases)
ws_s["A4"] = "P0"
ws_s["B4"] = sum(1 for c in cases if c["Priority"] == "P0")
ws_s["A5"] = "P1"
ws_s["B5"] = sum(1 for c in cases if c["Priority"] == "P1")
ws_s["A6"] = "P2"
ws_s["B6"] = sum(1 for c in cases if c["Priority"] == "P2")
ws_s["A8"] = "By Type"
ws_s["A8"].font = Font(bold=True)
types = {}
for c in cases:
    types[c["Type"]] = types.get(c["Type"], 0) + 1
r = 9
for t, n in sorted(types.items(), key=lambda x: -x[1]):
    ws_s[f"A{r}"] = t
    ws_s[f"B{r}"] = n
    r += 1
r += 1
ws_s[f"A{r}"] = "By Role"
ws_s[f"A{r}"].font = Font(bold=True)
r += 1
roles = {}
for c in cases:
    roles[c["Role"]] = roles.get(c["Role"], 0) + 1
for role, n in sorted(roles.items(), key=lambda x: -x[1]):
    ws_s[f"A{r}"] = role
    ws_s[f"B{r}"] = n
    r += 1
r += 2
ws_s[f"A{r}"] = "By Module"
ws_s[f"A{r}"].font = Font(bold=True)
r += 1
mods = {}
for c in cases:
    mods[c["Module"]] = mods.get(c["Module"], 0) + 1
for m, n in sorted(mods.items(), key=lambda x: -x[1]):
    ws_s[f"A{r}"] = m
    ws_s[f"B{r}"] = n
    r += 1

ws_s.column_dimensions["A"].width = 40
ws_s.column_dimensions["B"].width = 12

# Traceability
ws_t = wb.create_sheet("98_Coverage_Matrix", 1)
ws_t["A1"] = "Coverage Matrix (Module × Role × Types present)"
ws_t["A1"].font = Font(bold=True, size=12)
ws_t.append(["Module", "Admin", "Landlord", "Tenant", "Manager", "All", "P0 count", "Has Security?", "Has Negative?"])
for m, rows in sorted(groups.items()):
    def has_role(role):
        return "Y" if any(c["Role"] == role for c in rows) else ""
    ws_t.append([
        m,
        has_role("Admin"),
        has_role("Landlord"),
        has_role("Tenant"),
        has_role("Manager"),
        has_role("All"),
        sum(1 for c in rows if c["Priority"] == "P0"),
        "Y" if any(c["Type"] == "Security" for c in rows) else "",
        "Y" if any(c["Type"] == "Negative" for c in rows) else "",
    ])
for col in range(1, 10):
    ws_t.column_dimensions[get_column_letter(col)].width = 14
ws_t.column_dimensions["A"].width = 36

wb.save(OUT)

ASSUMPTIONS.write_text("""# RPMS Test Case Suite — Assumptions & Gaps

## Nguồn phân tích
- Source code RPMS (.NET 8 WinForms / BLL / DAL)
- `Docs/TongQuanDuAn_RPMS.docx`
- `Database/RPMS_Full.sql` + `DatabaseSchemaUpdater`
- Hành vi UI đã implement (Login, Contract Draft/Active, Meter tháng trước, Assignment theo Manager ID/Username…)

## Giả định
1. **RegisterForm** tạo user Active; role mặc định tùy UI (cần xác nhận Role mặc định khi đăng ký).
2. **BackupForm / BackupService**: Program.cs đăng ký nhưng file có thể thiếu trên disk — TC Backup ghi nhận “pass nếu có / fail rõ nếu thiếu”.
3. **Pay invoice notify Landlord**: có thể chưa implement notify chủ khi thanh toán — TC ghi “nếu implement”.
4. **Appointment past date**: validation phía UI có thể chưa chặn — TC exploratory/boundary.
5. **Email format validation**: có thể chỉ Unique, chưa regex — TC ghi nhận hành vi thực tế.
6. **Session timeout**: app desktop không có JWT timeout server; TC Logout thủ công thay cho timeout.
7. **XSS**: WinForms không execute HTML; TC security ghi nhận lưu plain text.
8. **Pagination**: nhiều list không phân trang server — TC Performance/Usability thay Pagination.
9. **Import**: không có import hàng loạt user/room — ngoài scope, không tạo TC Import trừ media upload.
10. **Direct URL bypass**: không có web URL; “Bypass UI” = không có menu + không resolve form trái role từ shell.

## Mâu thuẫn / rủi ro tài liệu
- README còn nói Admin “Phân công Manager” trong khi code đã chuyển sang **LandlordAssignmentForm**.
- CHECK Appointment: script SQL có `Completed`; EF config có thể khác — runtime DB là nguồn đúng.
- `IBackupService` đăng ký trong DI trong khi file implement có thể thiếu → build/runtime risk.

## Cách dùng file Excel
- Sheet `99_Summary`: tổng quan số lượng.
- Sheet `98_Coverage_Matrix`: bao phủ module × role.
- Sheet `00_All_TestCases`: full list + AutoFilter.
- Các sheet theo Module: thực thi theo đội.
- Cột **Actual Result / Status** để trống cho QA điền (Pass/Fail/Blocked).
- Priority: P0 blocker/smoke; P1 core; P2 nice-to-have.

## Khuyến nghị thực thi
1. Chạy hết **Smoke Suite** + **Regression Hotspots** trước mỗi build.
2. Chạy full P0 theo role trên môi trường SQLEXPRESS seed.
3. Automation Candidate=Yes ưu tiên API/service-level (xUnit + DB test), UI WinForms tự động hóa sau.
""", encoding="utf-8")

print("Wrote", OUT, "size", OUT.stat().st_size)
print("Wrote", ASSUMPTIONS)
